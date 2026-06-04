"""
conexia_bot.py — Bot Conexia SEROS (versión final simplificada)
================================================================
Estrategia: probar page + frame_locator("iframe").nth(0,1,2) en cada acción.
No asume dónde está el contenido — lo busca en los 4 contextos posibles.
"""

import asyncio, random, logging, zipfile, re, sys, json
from pathlib import Path
from datetime import date, datetime
from collections import defaultdict
from urllib.parse import urljoin

import openpyxl
from dateutil.relativedelta import relativedelta
from playwright.async_api import async_playwright, Page, TimeoutError as PwTimeout
from dotenv import load_dotenv
import os
load_dotenv()

_ant  = date.today() - relativedelta(months=1)
MESES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
         7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
MES   = MESES[_ant.month]
ANIO  = str(_ant.year)

URL      = "https://seros.conexia.com.ar:8443/WebPrestador/"
EXCEL    = Path(os.getenv("EXCEL_PATH","credenciales_seros.xlsx"))
DESDE    = int(os.getenv("DESDE", "0"))   # registro inicial 1-based (0=sin límite)
HASTA    = int(os.getenv("HASTA", "0"))   # registro final 1-based inclusivo (0=sin límite)
# Subcarpeta por fecha de ejecución: output/20260524_HHmm/
_run_ts  = datetime.now().strftime("%Y%m%d_%H%M")
SALIDA   = Path(os.getenv("OUTPUT_DIR","./output")) / _run_ts
HEADLESS = os.getenv("HEADLESS","false").lower() == "true"
T        = 25_000

SALIDA.mkdir(parents=True, exist_ok=True)
Path("logs").mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)-8s %(message)s",
    handlers=[logging.StreamHandler(),
              logging.FileHandler("logs/bot.log", encoding="utf-8")]
)
log = logging.getLogger("conexia")
log.info(f"Período: {MES} {ANIO}")


# ════════════════════════════════════════════════════════════════════════
# EXCEL
# ════════════════════════════════════════════════════════════════════════
_SIN = {
    "usuario":    ["usuario","user","login","username"],
    "clave":      ["clave","password","contrasena","contraseña","pass"],
    "profesional":["profesional","profesional a elegir","prestador","nombre"],
    "cuit":       ["cuit","cuit/cuil","cuil","nro cuit"],
}
def _norm(t):
    t = str(t).lower().strip()
    for a,b in [("á","a"),("é","e"),("í","i"),("ó","o"),("ú","u"),("ñ","n")]:
        t = t.replace(a,b)
    return t
def _col(hdrs,campo):
    for i,h in enumerate(hdrs):
        hn=_norm(h)
        if any(hn==_norm(s) or hn.startswith(_norm(s)) for s in _SIN.get(campo,[campo])):
            return i
    return None
def leer_excel(path, desde: int = 0, hasta: int = 0):
    if not path.exists(): log.error(f"No encontrado: {path}"); sys.exit(1)
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    hdrs,hrow=[],0
    for i,row in enumerate(ws.iter_rows(values_only=True),1):
        if sum(1 for c in row if c)>=2:
            hdrs=[str(c) if c else "" for c in row]; hrow=i; break
    col={c:_col(hdrs,c) for c in ("usuario","clave","profesional","cuit")}
    log.info(f"Columnas: {col}")
    grupos=defaultdict(list)
    for row in ws.iter_rows(min_row=hrow+1, values_only=True):
        if all(c is None or str(c).strip()=='' for c in row): continue
        def g(i): return str(row[i]).strip() if i is not None and i<len(row) and row[i] else ""
        u,c=g(col["usuario"]),g(col["clave"])
        if not u or not c: continue
        grupos[(u,c)].append({"nombre":g(col["profesional"]),"cuit":g(col["cuit"])})
    grupos_dict = dict(grupos)

    # Filtro de rango si se especificó (1-based)
    if desde > 0 or hasta > 0:
        todos_reg = [(k, p) for k, prests in grupos_dict.items() for p in prests]
        d = (desde - 1) if desde > 0 else 0
        h = hasta if hasta > 0 else len(todos_reg)
        filtrado: dict = {}
        for (k, p) in todos_reg[d:h]:
            filtrado.setdefault(k, []).append(p)
        tot = sum(len(v) for v in filtrado.values())
        log.info(f"Excel: {tot} registros (rango {desde}-{h} de {len(todos_reg)} total)")
        return filtrado

    # Filtro de rango si se especificó (1-based)
    if desde > 0 or hasta > 0:
        todos_reg = [(k, p) for k, prests in grupos_dict.items() for p in prests]
        d = (desde - 1) if desde > 0 else 0
        h = hasta if hasta > 0 else len(todos_reg)
        filtrado: dict = {}
        for (k, p) in todos_reg[d:h]:
            filtrado.setdefault(k, []).append(p)
        tot = sum(len(v) for v in filtrado.values())
        log.info(f"Excel: {tot} registros (rango {desde}-{h} de {len(todos_reg)} total)")
        return filtrado

    log.info(f"Excel: {len(grupos_dict)} usuarios, {sum(len(v) for v in grupos_dict.values())} registros")
    return grupos_dict


# ════════════════════════════════════════════════════════════════════════
# CONTEXTOS: page + iframes 0,1,2
# ════════════════════════════════════════════════════════════════════════
def contextos(page: Page):
    """Devuelve [page, iframe0, iframe1, iframe2] para probar en orden."""
    return [
        page,
        page.frame_locator("iframe").nth(0),
        page.frame_locator("iframe").nth(1),
        page.frame_locator("iframe").nth(2),
    ]

async def buscar_en_contextos(page: Page, selector: str, timeout_ms=3000):
    """
    Busca el selector en page y en los 3 primeros iframes.
    Retorna (contexto, locator) del primero que encuentre elementos visibles.
    """
    for ctx in contextos(page):
        try:
            loc = ctx.locator(selector)
            if await loc.first.is_visible(timeout=timeout_ms):
                return ctx, loc
        except Exception:
            continue
    return None, None

async def pausa(a=1.0,b=3.0): await asyncio.sleep(random.uniform(a,b))
def nom_seg(t,n=60): return re.sub(r"\s+","_",re.sub(r"[^\w\s-]","",str(t).strip())).upper()[:n]

def nom_archivo(t, n=90):
    """Nombre de archivo LEGIBLE: conserva espacios, puntos y acentos;
    solo quita los caracteres ilegales en un nombre de archivo."""
    t = re.sub(r'[\\/:*?"<>|\r\n\t]', "", str(t)).strip()
    return t[:n] or "SIN_NOMBRE"

# Orígenes del "Reporte de Facturación / Facturación" en el menú de Conexia.
# Cada uno: (fragmento del href del link, etiqueta, formato de descarga).
# - Prácticas se baja en PDF: su HTML no trae Nombre/Documento del afiliado,
#   pero el PDF del servidor sí los incluye.
# - Ambulatorio y Prácticas Esp. se bajan en HTML (ya traen todo).
# Nota: "Reportes → Facturación" usa el mismo endpoint que "Ambulatorio".
ORIGENES_REPORTE = [
    ("menuReporteTotWebFactuPres.do",         "Ambulatorio",    "html"),
    ("menuReporteTotWebFactuRadiologia.do",   "Prácticas",      "pdf"),
    ("menuReporteTotWebFactuEspecialista.do", "Prácticas Esp.", "html"),
]

def mk_dir(u,s):
    p=SALIDA/f"{datetime.now().strftime('%Y%m%d')}_{nom_seg(u+'_'+s)}"
    p.mkdir(parents=True,exist_ok=True); return p


# ════════════════════════════════════════════════════════════════════════
# DETECCIÓN DE PANTALLA
# ════════════════════════════════════════════════════════════════════════
async def que_veo(page: Page) -> str:
    """
    Lee texto de page.frames (frames reales) para detectar la pantalla actual.
    page.frames[0] = frameset raíz
    page.frames[1] = menú (init.do)
    page.frames[2] = contenido (cambia según lo que se cargó)
    """
    try:
        url = page.url.lower()
        txt = ""

        # Leer de todos los frames reales
        for fr in page.frames:
            try:
                if fr.url in ("about:blank", ""): continue
                t = (await fr.locator("body").inner_text(timeout=3_000)).lower()
                if t: txt += " " + t
            except Exception:
                pass

        if not txt:
            # Fallback a frame_locator si page.frames no dio nada
            for ctx in contextos(page):
                try:
                    t = (await ctx.locator("body").inner_text(timeout=2_000)).lower()
                    if t: txt += " " + t; break
                except Exception: pass

        if any(e in txt for e in ["sesión expirada","session expired"]): return "error"
        if "exportar" in txt and "consultar" in txt:    return "facturacion"
        if "seleccione el periodo" in txt:               return "facturacion"
        if "cambiar prestador" in txt:                  return "dashboard"
        if "bienvenido" in txt and "init" not in url:   return "dashboard"
        if "lugar de atenci" in txt:                    return "lugar"
        if "seleccione el prestador" in txt:            return "prestador"
        if "init.do" in url:                            return "prestador"
        if "nombre usuario" in txt or "ingresar" in txt: return "login"
        if "inválido" in txt or "invalido" in txt or "clave es inv" in txt:
            return "error_credenciales"
        # Último intento: "seleccione una opción" indica pantalla de selección
        if "seleccione una" in txt or "seleccione el" in txt: return "prestador"
        if txt.strip():
            log.warning(f"  que_veo=? texto ({len(txt)} chars): {txt[:300]}")
        else:
            log.warning("  que_veo=? texto vacío (página no cargó o frameset vacío)")
        return "?"
    except Exception as e:
        log.debug(f"que_veo: {e}"); return "?"


# ════════════════════════════════════════════════════════════════════════
# LOGIN
# ════════════════════════════════════════════════════════════════════════
async def login(page: Page, usuario: str, clave: str) -> bool:
    log.info(f"[{usuario}] Login...")
    for intento in range(3):
        try:
            await page.goto(URL, wait_until="domcontentloaded", timeout=T)

            # Esperar que el iframe de login esté listo (hasta 20s)
            log.info(f"[{usuario}] Esperando iframe... (intento {intento+1}/3)")
            iframe_listo = False
            for _ in range(20):
                try:
                    cnt = await page.frame_locator("iframe").first                               .locator("input[name='usuario']").count()
                    if cnt > 0:
                        iframe_listo = True
                        break
                except Exception:
                    pass
                await asyncio.sleep(1)

            if not iframe_listo:
                log.warning(f"[{usuario}] Iframe no cargó, reintentando...")
                await pausa(3,5)
                continue

            fr = page.frame_locator("iframe").first
            await fr.locator("input[name='usuario']").fill(usuario)
            await pausa(0.3,0.6)
            await fr.locator("input[name='clave']").fill(clave)
            await pausa(0.3,0.6)
            await fr.locator("#Continuar").click()

            for _ in range(25):
                await asyncio.sleep(1)
                url = page.url.lower()
                if "init.do" in url or "prestadorselected" in url: break
                try:
                    if await page.frame_locator("iframe").first                             .locator("input[name='usuario']").count() == 0: break
                except Exception: break

            await page.wait_for_load_state("networkidle", timeout=12_000)
            estado = await que_veo(page)
            log.info(f"[{usuario}] -> {estado} | {page.url}")
            if estado in ("prestador","lugar","dashboard","facturacion"):
                return True
            # Credenciales inválidas — no tiene sentido reintentar
            if estado == "error_credenciales":
                log.error(f"[{usuario}] Credenciales inválidas — verificar en el Excel.")
                return False
            # Si el iframe de login ya desapareció, asumir que entramos
            try:
                cnt = await page.frame_locator("iframe").first                           .locator("input[name='usuario']").count()
                if cnt == 0 and len(page.frames) > 1:
                    log.warning(f"[{usuario}] que_veo=? pero frames={len(page.frames)} → asumiendo OK")
                    return True
            except Exception:
                pass
            log.warning(f"[{usuario}] Estado inesperado: {estado}, reintentando...")
            await pausa(3,5)

        except Exception as e:
            log.warning(f"[{usuario}] intento {intento+1} error: {e}")
            await pausa(3,5)

    log.error(f"[{usuario}] Login fallido tras 3 intentos.")
    return False


async def opciones_select(page: Page) -> list[dict]:
    """Usa page.frames primero (frameset) y frame_locator como fallback."""
    ctx = None
    for fr in page.frames:
        try:
            if fr.url in ("about:blank",""): continue
            if await fr.locator("select").count() > 0:
                ctx = fr; break
        except Exception: pass
    if ctx is None:
        ctx_fl, _ = await buscar_en_contextos(page, "select", timeout_ms=4000)
        if ctx_fl is not None: ctx = ctx_fl
    if ctx is None: return []
    try:
        opts = await ctx.locator("select").first.locator("option").all()
        return [{"value": await o.get_attribute("value"),
                 "texto":(await o.inner_text()).strip()}
                for o in opts
                if (await o.get_attribute("value")) and
                   "seleccione" not in (await o.inner_text()).lower()]
    except Exception as e:
        log.debug(f"opciones_select: {e}"); return []


async def confirmar_opcion(page: Page, valor: str) -> bool:
    """
    Selecciona el valor en el select y hace clic en Confirmar.
    Busca en page.frames reales primero (más confiable que frame_locator).
    Verifica que la página navega al dashboard después de confirmar.
    """
    # Encontrar el frame con el select usando page.frames
    ctx = None
    for fr in page.frames:
        try:
            if fr.url in ("about:blank",""): continue
            if await fr.locator("select").count() > 0:
                ctx = fr
                log.info(f"  Select en frame: {fr.url[:60]}")
                break
        except Exception:
            pass
    # Fallback a frame_locator
    if ctx is None:
        ctx_fl, _ = await buscar_en_contextos(page, "select", timeout_ms=4000)
        if ctx_fl is None:
            log.error("  confirmar_opcion: select no encontrado"); return False
        ctx = ctx_fl

    try:
        # Seleccionar el valor
        await ctx.locator("select").first.select_option(value=valor)
        log.info(f"  Opción seleccionada: {valor}")
        await pausa(0.4,0.8)

        # Buscar el botón Confirmar en el mismo frame
        btn_conf = None
        for sel in ["input[value='Confirmar']", "button:has-text('Confirmar')",
                    "input[value='Confirmar ']", "a:has-text('Confirmar')"]:
            b = ctx.locator(sel)
            if await b.count() > 0:
                btn_conf = b
                log.info(f"  Confirmar encontrado: {sel}")
                break

        if btn_conf is None:
            # JS fallback
            log.info("  Buscando Confirmar via JS...")
            elementos = await ctx.evaluate("""
                () => Array.from(document.querySelectorAll('input,button,a'))
                     .map(e => (e.value||e.textContent.trim()).substring(0,20))
                     .filter(t => t.length > 0)
            """)
            log.info(f"  Elementos en frame: {elementos}")
            clicado = await ctx.evaluate("""
                () => { for(var e of document.querySelectorAll('input,button,a')){
                    var t=(e.value||e.textContent||'').trim();
                    if(t==='Confirmar'){e.click();return true;}}return false;}
            """)
            if not clicado:
                log.error("  Botón Confirmar no encontrado.")
                return False
            log.info("  Confirmar via JS ✓")
        else:
            await btn_conf.first.click()
            log.info("  Confirmar clickeado ✓")

        # Esperar que navegue al dashboard (más frames deben aparecer)
        for _ in range(15):
            await asyncio.sleep(1)
            estado = await que_veo(page)
            log.info(f"  Post-confirmar estado: {estado}")
            if estado in ("dashboard","facturacion","prestador"):
                return True
            # Verificar si aparecieron más frames (señal de navegación)
            if len(page.frames) > 2:
                return True

        log.error("  No se llegó al dashboard tras confirmar.")
        return False

    except Exception as e:
        log.error(f"  confirmar_opcion: {e}"); return False

async def sel_prestador(page: Page, cuit: str, nombre: str) -> bool:
    opts = await opciones_select(page)
    if not opts: log.error("  Sin opciones de prestador."); return False
    elegido = (next((o for o in opts if cuit and cuit in o["texto"]),None)
            or next((o for o in opts if nombre and
                     nombre.strip().split()[0].upper() in o["texto"].upper()),None)
            or (opts[0] if len(opts)==1 else None))
    if not elegido:
        log.error(f"  No encontrado CUIT={cuit}. Opciones:{[o['texto'] for o in opts]}")
        return False
    log.info(f"  → {elegido['texto']}")
    return await confirmar_opcion(page, elegido["value"])


# ════════════════════════════════════════════════════════════════════════
# NAVEGAR A FACTURACIÓN
# ════════════════════════════════════════════════════════════════════════
async def ir_facturacion(page: Page) -> bool:
    """
    1. Loguea todos los frames disponibles para diagnóstico
    2. En cada frame/contexto, expande TODAS las secciones colapsables del menú
    3. Busca y hace clic en "Reporte Facturación" o "Facturación"
    4. Verifica que el contenido del centro cambió a la pantalla de Facturación
    """
    log.info("  → Facturación...")

    # ── Diagnóstico: listar todos los frames disponibles ──────────────────
    frames_reales = page.frames
    log.info(f"  Frames disponibles ({len(frames_reales)}):")
    for i, fr in enumerate(frames_reales):
        try:
            cnt_links = await fr.locator("a").count()
            cnt_sels  = await fr.locator("select").count()
            log.info(f"    [{i}] links={cnt_links} selects={cnt_sels} url={fr.url[:80]}")
        except Exception:
            log.info(f"    [{i}] (no accesible) url={fr.url[:60]}")

    # ── Construir lista completa de contextos a probar ────────────────────
    # Incluye frames reales (page.frames) Y frame_locators (iframes CSS)
    todos_ctx = list(frames_reales) + [
        page.frame_locator("iframe").nth(0),
        page.frame_locator("iframe").nth(1),
        page.frame_locator("iframe").nth(2),
    ]

    # ── Para cada contexto: expandir todo y buscar Facturación ────────────
    for ctx in todos_ctx:
        try:
            # Verificar que el contexto tiene links (es el menú)
            cnt = await ctx.locator("a").count()
            if cnt == 0:
                continue

            log.info(f"  Procesando contexto con {cnt} links...")

            # Expandir TODAS las secciones colapsables (tienen > o ▼ o son toggleables)
                        # Solo headers padre expandibles.
            # IMPORTANTE: usar :text-is() (exacto) NO :has-text() (substring)
            # has-text('Prácticas') también matchea 'Solicitud de Prácticas'!
            HEADERS_PADRE = ["Ambulatorio","Prácticas","Prácticas Esp.","Reportes","Gestión","Documentos"]
            sel_headers = ", ".join(f"a:text-is(\'{h}\')" for h in HEADERS_PADRE)
            expandibles = ctx.locator(sel_headers)
            n_exp = await expandibles.count()
            log.info(f"  Expandiendo {n_exp} secciones del menú...")
            for i in range(n_exp):
                try:
                    lnk = expandibles.nth(i)
                    if await lnk.is_visible(timeout=1000):
                        txt = await lnk.inner_text()
                        log.info(f"    Expandiendo: {txt.strip()}")
                        await lnk.click()
                        await pausa(0.4, 0.7)
                except Exception:
                    pass

            # Ahora buscar Facturación o Reporte Facturación (puede estar visible ahora)
            for sel_fac in [
                "a:has-text('Reporte Facturación')",
                "a:has-text('Reporte de Facturación')",
                "a:has-text('Facturación')",
            ]:
                links_fac = ctx.locator(sel_fac)
                n_fac = await links_fac.count()
                log.info(f"  '{sel_fac}': {n_fac} encontrados")
                for j in range(n_fac):
                    try:
                        lnk = links_fac.nth(j)
                        if await lnk.is_visible(timeout=1500):
                            txt = await lnk.inner_text()
                            log.info(f"  ✓ Clic en: {txt.strip()}")
                            await lnk.click()
                            await pausa(1.5, 2.5)

                            # Esperar que el frame de contenido (frames[2]) cargue el formulario
                            for _ in range(10):
                                await asyncio.sleep(1)
                                # Verificar en frames reales si hay formulario de facturación
                                for fr in page.frames:
                                    try:
                                        if fr.url in ("about:blank",""): continue
                                        t = (await fr.locator("body").inner_text(timeout=2_000)).lower()
                                        if ("exportar" in t and "consultar" in t) or "seleccione el periodo" in t:
                                            log.info(f"  En Facturación ✓ (frame: {fr.url[:60]})")
                                            return True
                                    except Exception:
                                        pass
                            log.warning(f"  Clic hecho pero formulario no apareció")
                    except Exception as e:
                        log.debug(f"  {sel_fac}[{j}]: {e}")

        except Exception as e:
            log.debug(f"  contexto: {e}")
            continue

    log.error("  Facturación no encontrada en ningún contexto.")
    return False


async def ir_a_reporte(page: Page, href_frag: str, etiqueta: str) -> bool:
    """
    Navega al 'Reporte de Facturación / Facturación' de un origen concreto,
    identificado por un fragmento de su href (p.ej. 'menuReporteTotWebFactuRadiologia.do').
    Expande el menú, clickea ese link y espera a que aparezca el formulario.
    Devuelve False si ese origen no existe para el prestador (no es error).
    """
    log.info(f"  → Reporte [{etiqueta}] (href ~ {href_frag})...")

    todos_ctx = list(page.frames) + [
        page.frame_locator("iframe").nth(0),
        page.frame_locator("iframe").nth(1),
        page.frame_locator("iframe").nth(2),
    ]
    HEADERS_PADRE = ["Ambulatorio", "Prácticas", "Prácticas Esp.", "Reportes", "Gestión", "Documentos"]

    for ctx in todos_ctx:
        try:
            if await ctx.locator("a").count() == 0:
                continue

            # Expandir secciones del menú para que el link sea visible
            sel_headers = ", ".join(f"a:text-is(\'{h}\')" for h in HEADERS_PADRE)
            expandibles = ctx.locator(sel_headers)
            for i in range(await expandibles.count()):
                try:
                    lnk = expandibles.nth(i)
                    if await lnk.is_visible(timeout=800):
                        await lnk.click()
                        await pausa(0.3, 0.6)
                except Exception:
                    pass

            # Buscar el link por su href
            links = ctx.locator(f"a[href*='{href_frag}']")
            n = await links.count()
            if n == 0:
                continue

            clicado = False
            for j in range(n):
                try:
                    lk = links.nth(j)
                    if await lk.is_visible(timeout=1000):
                        await lk.click()
                        clicado = True
                        break
                except Exception:
                    pass
            # Si ninguno estaba visible, forzar el clic por JS sobre el href
            if not clicado:
                try:
                    clicado = await ctx.evaluate(
                        """(h) => { var a=document.querySelector("a[href*='"+h+"']");
                                    if(a){ a.click(); return true; } return false; }""",
                        href_frag,
                    )
                except Exception:
                    clicado = False
            if not clicado:
                continue

            await pausa(1.5, 2.5)

            # Esperar a que el formulario de facturación aparezca
            for _ in range(10):
                await asyncio.sleep(1)
                for fr in page.frames:
                    try:
                        if fr.url in ("about:blank", ""):
                            continue
                        t = (await fr.locator("body").inner_text(timeout=2_000)).lower()
                        if ("exportar" in t and "consultar" in t) or "seleccione el periodo" in t:
                            log.info(f"  En reporte [{etiqueta}] ✓")
                            return True
                    except Exception:
                        pass
            log.warning(f"  [{etiqueta}] clic hecho pero el formulario no apareció")
        except Exception as e:
            log.debug(f"  ir_a_reporte[{etiqueta}]: {e}")
            continue

    log.warning(f"  Reporte [{etiqueta}] no encontrado para este prestador.")
    return False


# ════════════════════════════════════════════════════════════════════════
# DESCARGAR HTML
# ════════════════════════════════════════════════════════════════════════
def _js_form_fill() -> str:
    """JS que selecciona Mes, Año y el radio del formulario de Facturación."""
    return f"""
    () => {{
        var I = {{s:0, mes:false, anio:false, r:0, ok:false, err:[]}};
        var sels = document.querySelectorAll('select');
        I.s = sels.length;
        var sM=null,sA=null;
        for(var s of sels){{
            var ts=Array.from(s.options).map(o=>o.text.trim());
            if(ts.includes('{MES}'))  sM=s;
            if(ts.includes('{ANIO}')) sA=s;
        }}
        if(sM){{ for(var i=0;i<sM.options.length;i++){{
            if(sM.options[i].text.trim()==='{MES}'){{
                sM.selectedIndex=i;
                sM.dispatchEvent(new Event('change',{{bubbles:true}}));
                I.mes=true; break;}}}} }}
        else I.err.push('Mes');
        if(sA){{ for(var i=0;i<sA.options.length;i++){{
            if(sA.options[i].text.trim()==='{ANIO}'){{
                sA.selectedIndex=i;
                sA.dispatchEvent(new Event('change',{{bubbles:true}}));
                I.anio=true; break;}}}} }}
        else I.err.push('Año');
        var rs=document.querySelectorAll('input[type="radio"]');
        I.r=rs.length;
        if(rs.length>1){{
            for(var r of rs) r.checked=false;
            rs[1].checked=true; rs[1].click();
            rs[1].dispatchEvent(new Event('change',{{bubbles:true}}));
            I.ok=true;
        }} else if(rs.length==1){{ rs[0].checked=true; rs[0].click(); I.ok=true; }}
        else I.err.push('Radios');
        return I;
    }}
    """


async def descargar_pdf(page: Page, dest: Path, nom: str) -> Path | None:
    """
    Llena el formulario y descarga el PDF del reporte clickeando el botón con la
    imagen 'type_file_pdf.gif'. El PDF del servidor trae los datos completos
    (Nombre y Documento del afiliado), que el HTML de Prácticas no incluye.
    """
    log.info(f"  ↓ PDF — {MES} {ANIO}")
    try:
        await page.wait_for_load_state("networkidle", timeout=10_000)

        ctx = None
        for fr in page.frames:
            try:
                if fr.url in ("about:blank", ""):
                    continue
                if await fr.locator("select").count() >= 1:
                    ctx = fr; break
            except Exception:
                pass
        if ctx is None:
            ctx, _ = await buscar_en_contextos(page, "select", timeout_ms=5000)
        if ctx is None:
            log.error("  PDF: sin selects en ningún contexto."); return None

        fi = await ctx.evaluate(_js_form_fill())
        log.info(f"  Form JS (pdf): {fi}")
        if not fi["mes"] or not fi["anio"]:
            log.error(f"  PDF: fecha no seleccionada: {fi['err']}"); return None
        await pausa(0.3, 0.6)

        # Ubicar el frame y el elemento del botón PDF (img type_file_pdf.gif)
        fr_pdf, loc_pdf = None, None
        for fr in page.frames:
            try:
                loc = fr.locator("img[src*='type_file_pdf']")
                if await loc.count() > 0:
                    fr_pdf, loc_pdf = fr, loc.first
                    break
            except Exception:
                pass

        if loc_pdf is None:
            log.warning("  PDF: botón type_file_pdf no encontrado. Diagnóstico:")
            await _diag_exportar(page)
            return None

        destino = dest / f"{nom}.pdf"
        k = 2
        while destino.exists():
            destino = dest / f"{nom}_{k}.pdf"; k += 1

        def _guardar(b: bytes) -> bool:
            if b[:4] == b"%PDF":
                destino.write_bytes(b); return True
            if b[:2] == b"PK":   # vino comprimido
                import io
                with zipfile.ZipFile(io.BytesIO(b)) as zf:
                    pdfs = [a for a in zf.namelist() if a.lower().endswith(".pdf")]
                    if pdfs:
                        destino.write_bytes(zf.read(pdfs[0])); return True
            return False

        # ── Plan A: leer el href/onclick real del botón y bajar el ZIP directo
        #    por su URL (con la misma sesión). Es lo más robusto. ──────────────
        info = await fr_pdf.evaluate("""() => {
            var img=document.querySelector("img[src*='type_file_pdf']");
            if(!img) return null;
            var a=img.closest('a');
            return {
                href:   a ? a.getAttribute('href') : null,
                onclick:(a ? a.getAttribute('onclick') : null) || img.getAttribute('onclick'),
                base:   document.baseURI
            };
        }""")
        log.info(f"  PDF botón → {info}")

        urls_candidatas = []
        if info:
            base = info.get("base") or fr_pdf.url
            href = (info.get("href") or "").strip()
            oc   = (info.get("onclick") or "")
            if href and not href.lower().startswith("javascript:") and href != "#":
                urls_candidatas.append(urljoin(base, href))
            # URLs dentro del onclick (window.open('...'), location='...', etc.)
            for m in re.findall(r"""['"]([^'"]+\.(?:do|zip|pdf)(?:\?[^'"]*)?)['"]""", oc):
                urls_candidatas.append(urljoin(base, m))

        for u in urls_candidatas:
            try:
                resp = await page.context.request.get(u)
                if resp.ok and _guardar(await resp.body()):
                    log.info(f"  ✓ {destino.name} (url directa)")
                    return destino
            except Exception:
                pass

        # ── Plan B: un solo clic. La descarga (ZIP con el PDF) la dispara una
        #    ventana paralela, así que adjuntamos el oyente de 'download' tanto a
        #    la página principal como a CUALQUIER ventana nueva apenas aparece.
        #    Así no importa el timing ni en qué página caiga la descarga. ──────
        holder = {}
        ev = asyncio.Event()

        def _on_download(d):
            if "dl" not in holder:
                holder["dl"] = d
                ev.set()

        def _on_page(p):
            try: p.on("download", _on_download)
            except Exception: pass
            holder.setdefault("popups", []).append(p)

        page.on("download", _on_download)
        page.context.on("page", _on_page)
        try:
            try:
                await loc_pdf.click()
            except Exception:
                await loc_pdf.click(force=True)
            try:
                await asyncio.wait_for(ev.wait(), timeout=35)
            except asyncio.TimeoutError:
                pass
        finally:
            try: page.remove_listener("download", _on_download)
            except Exception: pass
            try: page.context.remove_listener("page", _on_page)
            except Exception: pass

        # ¿Capturamos la descarga (en la página o en la ventana paralela)?
        if "dl" in holder:
            try:
                d = holder["dl"]
                tmp = dest / (d.suggested_filename or f"{nom}.zip")
                await d.save_as(str(tmp))
                ok = _guardar(tmp.read_bytes())   # extrae el PDF del ZIP
                try: tmp.unlink(missing_ok=True)
                except Exception: pass
                # cerrar ventanas paralelas abiertas
                for p in holder.get("popups", []):
                    try: await p.close()
                    except Exception: pass
                if ok:
                    log.info(f"  ✓ {destino.name}"); return destino
            except Exception as e:
                log.error(f"  PDF: error guardando descarga: {e}")

        # Plan B: alguna ventana paralela quedó con la URL del PDF/ZIP → bajarla
        for p in holder.get("popups", []):
            try:
                u = p.url
                if u and not u.startswith("about:"):
                    resp = await page.context.request.get(u)
                    if _guardar(await resp.body()):
                        log.info(f"  ✓ {destino.name} (url ventana)")
                        for pp in holder.get("popups", []):
                            try: await pp.close()
                            except Exception: pass
                        return destino
            except Exception:
                pass
        for p in holder.get("popups", []):
            try: await p.close()
            except Exception: pass

        log.warning("  PDF: el botón existe pero no se pudo capturar el archivo. Diagnóstico:")
        await _diag_exportar(page)
        return None

    except (PwTimeout, asyncio.TimeoutError):
        log.error("  PDF: timeout."); return None
    except Exception as e:
        log.error(f"  descargar_pdf: {e}"); return None


async def _diag_exportar(page: Page):
    """Loguea los elementos de exportación (imgs/links/onclick) de cada frame,
    para diagnosticar cómo se dispara la descarga del PDF en Conexia."""
    for i, fr in enumerate(page.frames):
        try:
            info = await fr.evaluate("""() => {
                const out = {imgs:[], links:[], onclicks:[]};
                document.querySelectorAll('img').forEach(im => {
                    const s = im.getAttribute('src')||'';
                    if (s.toLowerCase().includes('pdf') || s.toLowerCase().includes('type_file'))
                        out.imgs.push({src:s, parent: (im.parentElement?im.parentElement.tagName:''),
                                       href: (im.closest('a')?im.closest('a').getAttribute('href'):''),
                                       onclick: (im.closest('a')?im.closest('a').getAttribute('onclick'):'')});
                });
                document.querySelectorAll('a').forEach(a => {
                    const h=(a.getAttribute('href')||''), oc=(a.getAttribute('onclick')||'');
                    if ((h+oc).toLowerCase().includes('pdf')) out.links.push({href:h, onclick:oc, txt:a.textContent.trim().slice(0,20)});
                });
                document.querySelectorAll('[onclick]').forEach(e => {
                    const oc=(e.getAttribute('onclick')||'');
                    if (oc.toLowerCase().includes('pdf')) out.onclicks.push({tag:e.tagName, onclick:oc.slice(0,80)});
                });
                return out;
            }""")
            if info["imgs"] or info["links"] or info["onclicks"]:
                log.info(f"    [diag frame {i}] {info}")
        except Exception:
            pass


async def descargar_html(page: Page, dest: Path, nom: str) -> Path | None:
    log.info(f"  ↓ HTML — {MES} {ANIO}")
    try:
        await page.wait_for_load_state("networkidle", timeout=10_000)

        # Buscar el frame real que tiene los selects del formulario
        ctx = None
        # Primero buscar en page.frames (más confiable)
        for fr in page.frames:
            try:
                if fr.url in ("about:blank",""): continue
                cnt = await fr.locator("select").count()
                if cnt >= 1:
                    log.info(f"  Formulario en frame real: {fr.url[:70]}")
                    ctx = fr
                    break
            except Exception:
                pass
        # Fallback a frame_locator si no encontramos con frames reales
        if ctx is None:
            ctx, _ = await buscar_en_contextos(page, "select", timeout_ms=5000)
        if ctx is None:
            log.error("  Sin selects en ningún contexto.")
            return None

        log.info(f"  Contexto formulario: {type(ctx).__name__}")

        # Seleccionar Mes, Año y radio via JavaScript
        fi = await ctx.evaluate(_js_form_fill())
        log.info(f"  Form JS: {fi}")

        if not fi["mes"] or not fi["anio"]:
            log.error(f"  Fecha no seleccionada: {fi['err']}")
            return None

        await pausa(0.3,0.5)

        # Buscar botón Consultar — usar el mismo frame del formulario (ctx)
        # Intentar múltiples selectores por si el value tiene variaciones
        btn_selectors = [
            "input[value='Consultar']",
            "input[value*='Consultar']",   # cubre  Consultar 
            "button:has-text('Consultar')",
            "input[type='button'][value*='Consultar']",
            "input[type='submit'][value*='Consultar']",
            "a:has-text('Consultar')",
        ]
        loc_btn = None
        for sel in btn_selectors:
            try:
                b = ctx.locator(sel)
                if await b.count() > 0:
                    loc_btn = b
                    log.info(f"  Consultar encontrado con: {sel}")
                    break
            except Exception:
                pass

        # Si no encontramos con selectores, usar JS para hacer clic
        if loc_btn is None:
            log.info("  Consultar no encontrado con selectores, intentando JS...")
            # Loguear todos los inputs/buttons del frame para diagnóstico
            info_btns = await ctx.evaluate("""
                () => Array.from(document.querySelectorAll('input,button,a'))
                    .filter(e => e.offsetParent !== null)  // solo visibles
                    .map(e => e.tagName + '|' + (e.value||e.textContent.trim()).substring(0,30))
            """)
            log.info(f"  Elementos visibles en frame: {info_btns}")

            clicado = await ctx.evaluate("""
                () => {
                    var todos = document.querySelectorAll('input,button,a');
                    for(var e of todos){
                        var txt = (e.value || e.textContent || '').trim().toLowerCase();
                        if(txt === 'consultar' || txt.includes('consultar')){
                            e.click(); return true;
                        }
                    }
                    return false;
                }
            """)
            if not clicado:
                log.error("  Botón Consultar no encontrado en ninguna forma.")
                return None
            log.info("  Consultar via JS ✓")
            # Esperar la descarga
            try:
                async with page.expect_download(timeout=T) as dl_info:
                    pass  # el clic ya fue disparado por JS
            except Exception:
                pass
        else:
            async with page.expect_download(timeout=T) as dl_info:
                await loc_btn.first.click()

        dl = await dl_info.value
        log.info(f"  ZIP: {dl.suggested_filename}")
        zp = dest/dl.suggested_filename
        await dl.save_as(str(zp))
        await pausa(0.3,0.5)
        return extraer_zip(zp, dest, nom)

    except (PwTimeout, asyncio.TimeoutError):
        log.error(f"  Timeout — ¿sin datos para {MES} {ANIO}?")
        return None
    except Exception as e:
        log.error(f"  descargar_html: {e}"); return None


def extraer_zip(zp: Path, dest: Path, nom: str) -> Path | None:
    try:
        with zipfile.ZipFile(zp,"r") as zf:
            archivos=zf.namelist(); log.info(f"  ZIP: {archivos}")
            t=(next((a for a in archivos if a.lower().endswith(".html")),None)
            or next((a for a in archivos if a.lower().endswith(".htm")),None)
            or (archivos[0] if archivos else None))
            if not t: log.warning("  ZIP vacío."); return None
            s=dest/f"{nom}.html"
            # Evitar sobrescribir si ya existe (p.ej. mismo profesional/origen
            # en otro lugar de atención): agregar sufijo numérico.
            k=2
            while s.exists():
                s=dest/f"{nom}_{k}.html"; k+=1
            s.write_bytes(zf.read(t))
            log.info(f"  ✓ {s.name}")
            try: zp.unlink(missing_ok=True)
            except Exception: pass
            return s
    except Exception as e:
        log.error(f"  extraer_zip: {e}"); return None


# ════════════════════════════════════════════════════════════════════════
# HTML → PDF
# ════════════════════════════════════════════════════════════════════════
async def html_a_pdf(html: Path, pw) -> Path | None:
    pdf=html.with_suffix(".pdf")
    try:
        br=await pw.chromium.launch(headless=True)
        pg=await br.new_page()
        await pg.goto(html.resolve().as_uri(), wait_until="networkidle", timeout=15_000)
        await pg.pdf(path=str(pdf), format="A4", print_background=True,
                     margin={"top":"10mm","bottom":"10mm","left":"10mm","right":"10mm"})
        await br.close(); log.info(f"  PDF ✓ {pdf.name}"); return pdf
    except Exception as e:
        log.error(f"  html_a_pdf: {e}"); return None


# ════════════════════════════════════════════════════════════════════════
# VOLVER AL INICIO
# ════════════════════════════════════════════════════════════════════════

# ════════════════════════════════════════════════════════════════════════
# MOTOR PRINCIPAL: login fresco por fila del Excel
# ════════════════════════════════════════════════════════════════════════

async def extraer_reporte(page, usuario: str, nombre: str, cuit: str,
                          sufijo_lugar: str = "") -> list:
    """
    Con sesión activa en el dashboard, extrae el 'Reporte de Facturación /
    Facturación' de CADA origen (Ambulatorio, Prácticas, Prácticas Esp.).
    Guarda un HTML/PDF por origen, nombrado: '{Nombre profesional}_{Origen}'.
    Devuelve una lista de resultados (uno por origen).
    """
    dest = mk_dir(usuario, nom_seg(nombre))
    base = nom_archivo(nombre)
    if sufijo_lugar:
        base = f"{base}_{nom_archivo(sufijo_lugar, 25)}"

    resultados = []
    for href_frag, origen, formato in ORIGENES_REPORTE:
        r = {"usuario": usuario, "profesional": nombre, "cuit": cuit,
             "origen": origen, "html": None, "pdf": None,
             "estado": "sin_datos", "detalle": ""}
        try:
            encontrado = await ir_a_reporte(page, href_frag, origen)
        except Exception as e:
            log.error(f"  ir_a_reporte[{origen}]: {e}")
            encontrado = False

        if not encontrado:
            r["detalle"] = f"Origen {origen} no disponible para este prestador"
            resultados.append(r)
            continue

        # Nombre del archivo: 'Nombre profesional_Origen'
        nom_base = f"{base}_{origen}"
        archivo = None
        if formato == "pdf":
            archivo = await descargar_pdf(page, dest, nom_base)
            if archivo is None:
                # Fallback: si el PDF falla, al menos bajar el HTML
                log.warning(f"  {origen}: PDF falló, intento HTML como respaldo.")
                archivo = await descargar_html(page, dest, nom_base)
        else:
            archivo = await descargar_html(page, dest, nom_base)

        # Se guarda la ruta en 'html' (campo de archivo genérico); el parser
        # detecta por extensión si es .pdf o .html.
        r["html"]    = str(archivo) if archivo else None
        r["estado"]  = "ok" if archivo else "sin_datos"
        r["detalle"] = "" if archivo else f"Sin datos {origen} {MES} {ANIO}"
        resultados.append(r)
        await pausa(1.0, 2.0)

    # Si NINGÚN origen tuvo datos, marcar un resultado representativo
    if not any(r["estado"] == "ok" for r in resultados):
        log.warning(f"  {nombre}: ningún reporte de facturación con datos.")
    return resultados


async def login_y_procesar(ctx, usuario: str, clave: str,
                            prof: dict, sufijo: str = "") -> list:
    """
    Abre página nueva, hace login fresco, selecciona el prestador/lugar
    indicado y extrae el reporte. Cierra la página al terminar.
    Retorna lista de resultados (puede ser más de uno si hay múltiples lugares).
    """
    nombre = prof["nombre"]
    cuit   = prof["cuit"]
    resultados = []

    page = await ctx.new_page()
    await page.add_init_script(
        "Object.defineProperty(navigator,'webdriver',{get:()=>undefined});"
    )

    def E(est, det):
        return {"usuario":usuario,"profesional":nombre,"cuit":cuit,
                "html":None,"pdf":None,"estado":est,"detalle":det}

    try:
        # ── Timeout máximo por usuario: 8 minutos ─────────────────────────────
        async def _procesar_con_timeout():
            ok = await login(page, usuario, clave)
            return ok

        try:
            ok = await asyncio.wait_for(login(page, usuario, clave), timeout=480)
        except asyncio.TimeoutError:
            log.error(f"  TIMEOUT 8min — {usuario} se canceló automáticamente")
            resultados.append(E("error_timeout", "Superó 8 minutos — se canceló"))
            await page.close()
            return resultados
        if not ok:
            resultados.append(E("error_login","Login fallido"))
            await page.close(); return resultados

        estado = await que_veo(page)
        log.info(f"  Post-login: {estado}")

        # ── PRESTADOR (dropdown de selección) ────────────────────────────────
        if estado == "prestador":
            ok = await sel_prestador(page, cuit, nombre)
            if not ok:
                resultados.append(E("error_seleccion", f"CUIT {cuit} no encontrado"))
                await page.close(); return resultados
            await pausa(0.5,1.0)
            estado = await que_veo(page)

        # ── LUGAR DE ATENCIÓN ────────────────────────────────────────────────
        if estado == "lugar":
            if sufijo:
                # Ya sabemos qué lugar elegir (valor del select)
                ok = await confirmar_opcion(page, sufijo)
                if not ok:
                    resultados.append(E("error_lugar", f"No se pudo seleccionar lugar {sufijo}"))
                    await page.close(); return resultados
                estado = await que_veo(page)
            else:
                # Primera visita: descubrir todos los lugares y procesarlos
                # con logins frescos
                lugares = await opciones_select(page)
                log.info(f"  Lugares detectados: {[l['texto'] for l in lugares]}")
                await page.close()  # cerramos esta página

                for lugar in lugares:
                    suf   = re.sub(r"\[.*?\]","",lugar["texto"]).strip()[:25]
                    na    = nom_seg(f"{nombre}_{suf}")
                    log.info(f"  → Lugar: {lugar['texto']}")
                    # Login fresco para cada lugar
                    res_lugar = await login_y_procesar(
                        ctx, usuario, clave, prof, sufijo=lugar["value"]
                    )
                    # Ajustar nombre de carpeta con sufijo del lugar
                    for r in res_lugar:
                        if r.get("html"):
                            r["lugar"] = lugar["texto"]
                        resultados.append(r)
                    await pausa(2,4)

                return resultados

        # ── DASHBOARD (prestador ya seleccionado) ────────────────────────────
        if estado in ("dashboard","facturacion"):
            rs = await extraer_reporte(page, usuario, nombre, cuit, sufijo_lugar=sufijo or "")
            for r in rs:
                if sufijo:
                    r["lugar"] = sufijo
                resultados.append(r)
        else:
            resultados.append(E(f"error_{estado}", f"Estado inesperado: {estado}"))

    except Exception as e:
        log.error(f"  login_y_procesar error: {e}")
        resultados.append(E("error_inesperado", str(e)))
    finally:
        try:
            await page.close()
        except Exception:
            pass

    return resultados


async def procesar_usuario(ctx, usuario: str, clave: str,
                            prestadores: list) -> list:
    """
    Procesa cada fila del Excel con un LOGIN FRESCO por fila.
    Evita todos los problemas de session loss al navegar entre prestadores.
    """
    todos = []
    log.info(f"\n{'═'*56}\nUSUARIO: {usuario}  ({len(prestadores)} registros)\n{'═'*56}")

    for idx, prof in enumerate(prestadores, 1):
        nombre = prof["nombre"]
        cuit   = prof["cuit"]
        log.info(f"\n  [{idx}/{len(prestadores)}] {nombre} | CUIT={cuit}")

        resultados = await login_y_procesar(ctx, usuario, clave, prof)
        todos.extend(resultados)

        # Pausa entre filas del mismo usuario
        if idx < len(prestadores):
            espera = random.uniform(5, 12)
            log.info(f"  Pausa {espera:.0f}s antes del siguiente registro...")
            await asyncio.sleep(espera)

    return todos


# ════════════════════════════════════════════════════════════════════════
# REPORTE
# ════════════════════════════════════════════════════════════════════════
def guardar_reporte(todos):
    ts=datetime.now().strftime("%Y%m%d_%H%M%S")
    fall=[r for r in todos if r.get("estado")!="ok"]
    (SALIDA/f"reporte_{ts}.json").write_text(
        json.dumps(todos,ensure_ascii=False,indent=2),encoding="utf-8")
    ls=[f"FALLIDOS — {MES} {ANIO}","─"*60,
        f"Total:{len(todos)} OK:{len(todos)-len(fall)} Fallidos:{len(fall)}","─"*60]
    for r in fall:
        ls.append(f"✗ {r['usuario']:15}|{r['profesional'][:35]:35}|{r['estado']}")
        if r.get("detalle"): ls.append(f"  └ {r['detalle']}")
    (SALIDA/f"FALLIDOS_{ts}.txt").write_text("\n".join(ls),encoding="utf-8")
    return fall


# ════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════
# ════════════════════════════════════════════════════════════════════════
# CONSOLIDACIÓN HTML → EXCEL
# ════════════════════════════════════════════════════════════════════════

def parsear_html(html_path: Path, nombre_profesional: str, cuit: str) -> dict:
    """
    Parsea el HTML de Conexia SEROS.
    Estructura real: los datos están como <td> individuales consecutivos en la tabla principal.
    Cada registro tiene exactamente 16 tokens: afiliado, nombre, codigo, desc, fecha, rrn,
    manual, nro_orden, cant, anticipo, saldo, mutual, total, imp_gastos, imp_honor, terminal.
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        log.error("pip install beautifulsoup4 lxml")
        return {"detalle": [], "totalizador": []}

    try:
        contenido = html_path.read_bytes()
        texto = ""
        for enc in ("utf-8","latin-1","cp1252"):
            try: texto = contenido.decode(enc); break
            except: pass
        if not texto:
            return {"detalle": [], "totalizador": []}

        soup = BeautifulSoup(texto, "lxml")
        # Tomar la tabla principal (índice 0, la más grande)
        tablas = soup.find_all("table")
        if not tablas:
            return {"detalle": [], "totalizador": []}

        tabla = tablas[0]
        # Extraer todos los textos no vacíos de los <td>
        tokens = []
        for td in tabla.find_all("td"):
            v = td.get_text(" ", strip=True).replace("\xa0"," ").strip()
            if v:
                tokens.append(v)

        if not tokens:
            return {"detalle": [], "totalizador": []}

        # Regex para detectar inicio de fila de detalle:
        # número de afiliado = 8 dígitos seguido de un nombre
        import re

        FECHA_RE   = re.compile(r"^\d{2}/\d{2}/\d{4}$")
        CODIGO_RE  = re.compile(r"^[A-Z]\d{7}$")   # ej: A4232010
        NUM_RE     = re.compile(r"^\d+\.?\d*$")
        AFIL_RE    = re.compile(r"^\d{8,10}$")
        WEBP_RE    = re.compile(r"^WEBP\d+$")
        TOTFACT_RE = re.compile(r"^[A-Z]\d{7}$")   # código PPM para totalizador

        detalle = []
        totalizador = []

        i = 0
        n = len(tokens)
        while i < n:
            t = tokens[i]

            # ── Detectar fila de DETALLE ──────────────────────────────────────
            # Patrón: AFIL_RE seguido de nombre (texto), código PPM, descripción, fecha
            if (AFIL_RE.match(t) and
                i+4 < n and
                FECHA_RE.match(tokens[i+4] if i+4 < n else "") and
                CODIGO_RE.match(tokens[i+2] if i+2 < n else "")):

                # Extraer los 16 campos
                def g(offset, default=""):
                    return tokens[i+offset] if i+offset < n else default
                def gn(offset):
                    try: return float(g(offset).replace(",","."))
                    except: return 0.0

                # Estructura exacta de 15 tokens por fila:
                # [0]afil [1]nombre [2]cod [3]desc [4]fecha [5]rrn [6]manual
                # [7]cant [8]anticipo [9]saldo [10]mutual [11]total [12]imp_gastos [13]imp_honor [14]terminal
                detalle.append({
                    "N° Afiliado":     g(0),
                    "Nombre Afiliado": g(1),
                    "Código PPM":      g(2),
                    "Descripción PPM": g(3),
                    "Fecha":           g(4),
                    "RRN":             g(5),
                    "Manual":          g(6),
                    "Nro Orden":       "",
                    "Cant":            gn(7),
                    "Anticipo":        gn(8),
                    "Saldo":           gn(9),
                    "Mutual":          gn(10),
                    "Total Facturado": gn(11),
                    "Imp Gastos":      gn(12),
                    "Imp Honor":       gn(13),
                    "Terminal":        g(14) if WEBP_RE.match(g(14)) else "",
                    "Profesional":     nombre_profesional,
                    "CUIT":            cuit,
                })
                i += 15
                continue

            # ── Detectar fila de TOTALIZADOR ──────────────────────────────────
            # Patrón: código PPM (A4232010) seguido de descripción y números
            if (CODIGO_RE.match(t) and
                i+7 < n and
                NUM_RE.match(tokens[i+2] if i+2 < n else "")):
                def g(offset, default=""):
                    return tokens[i+offset] if i+offset < n else default
                def gn(offset):
                    try: return float(g(offset).replace(",","."))
                    except: return 0.0

                totalizador.append({
                    "Código PPM":      g(0),
                    "Descripción PPM": g(1),
                    "Cantidad":        gn(2),
                    "Anticipo":        gn(3),
                    "Saldo":           gn(4),
                    "Mutual":          gn(5),
                    "Total Facturado": gn(6),
                    "Imp Gastos":      gn(7),
                    "Imp Honor":       gn(8),
                    "Profesional":     nombre_profesional,
                    "CUIT":            cuit,
                })
                i += 9
                continue

            i += 1

        log.info(f"  Parseado: {len(detalle)} filas detalle, {len(totalizador)} totalizador")
        return {"detalle": detalle, "totalizador": totalizador}

    except Exception as e:
        log.error(f"  parsear_html({html_path.name}): {e}")
        import traceback; log.debug(traceback.format_exc())
        return {"detalle": [], "totalizador": []}


# ════════════════════════════════════════════════════════════════════════
# PARSERS POR ORIGEN → ESQUEMA UNIFICADO (nivel detalle por prestación)
# ════════════════════════════════════════════════════════════════════════

# Columnas del Excel unificado. Los tres orígenes vuelcan acá a nivel
# detalle; lo que un reporte no tenga queda en blanco o en 0.
COLS_UNIF = [
    "Origen", "Profesional", "CUIT",
    "RRN", "Nro/Tipo Doc", "Nombre Afiliado",
    "Código PPM", "Descripción PPM", "Ámbito/Compl.", "Fecha",
    "Cant.", "Recargo Práctica", "Saldo", "Anticipo", "Total Fact.",
    "Honor.", "Gasto", "Coseguro", "Forma Pago", "Tipo", "Excep.",
]

import re as _re

_RRN9   = _re.compile(r"^\d{9}$")
_COD6   = _re.compile(r"^\d{6}$")
_DT     = _re.compile(r"^\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2}$")
_DOC    = _re.compile(r"^\d{2}-\d+$")
_NUMTOK = _re.compile(r"^\d[\d.,]*\.?$")


def _num(s):
    """Texto numérico → float (tolerante; '', '---' → 0). Maneja '65107.'."""
    s = str(s).strip().rstrip(".")
    if not s or s == "---":
        return 0.0
    try:
        return float(s.replace(",", ""))
    except Exception:
        return 0.0


def _fila_unif(origen, prof, cuit, **kw):
    base = {c: "" for c in COLS_UNIF}
    base["Origen"] = origen
    base["Profesional"] = prof
    base["CUIT"] = cuit
    base.update(kw)
    return base


def _tokens_html(html_path: Path) -> list:
    """Celdas no vacías del HTML, descartando la mega-celda concatenada."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        log.error("pip install beautifulsoup4 lxml")
        return []
    raw = html_path.read_bytes()
    txt = ""
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            txt = raw.decode(enc); break
        except Exception:
            pass
    if not txt:
        return []
    soup = BeautifulSoup(txt, "lxml")
    out = []
    for td in soup.find_all("td"):
        v = td.get_text(" ", strip=True).replace("\xa0", " ").strip()
        if v and len(v) < 200:        # descarta la mega-celda con todo el reporte
            out.append(v)
    return out


def _parse_ambulatorio(html_path, prof, cuit, origen):
    """Reusa el parser histórico de Ambulatorio y lo mapea al esquema unificado."""
    datos = parsear_html(html_path, prof, cuit)   # {'detalle':[...], 'totalizador':[...]}
    filas = []
    for d in datos.get("detalle", []):
        filas.append(_fila_unif(
            origen, prof, cuit,
            RRN=d.get("RRN", ""),
            **{"Nro/Tipo Doc":    d.get("N° Afiliado", ""),
               "Nombre Afiliado": d.get("Nombre Afiliado", ""),
               "Código PPM":      d.get("Código PPM", ""),
               "Descripción PPM": d.get("Descripción PPM", ""),
               "Fecha":           d.get("Fecha", ""),
               "Cant.":           d.get("Cant", 0),
               "Saldo":           d.get("Saldo", 0),
               "Anticipo":        d.get("Anticipo", 0),
               "Total Fact.":     d.get("Total Facturado", 0),
               "Honor.":          d.get("Imp Honor", 0),
               "Gasto":           d.get("Imp Gastos", 0)}))
    return filas


def _parse_practicas(html_path, prof, cuit, origen):
    """
    Detalle de Prácticas. Fila: RRN, Código(6), Ámbito(1 letra), Horario,
    Recargo, Anticipo, Total Fact., Honor., Forma Pago, Tipo.
    Saldo = Total − Anticipo; Gasto = Total − Honor; Coseguro = Anticipo
    (derivaciones confirmadas contra los totales del reporte).
    """
    t = _tokens_html(html_path)
    filas = []; i = 0; n = len(t)
    while i < n - 9:
        if (_RRN9.match(t[i]) and _COD6.match(t[i+1]) and
                len(t[i+2]) == 1 and t[i+2].isalpha() and _DT.match(t[i+3])):
            rec, ant = _num(t[i+4]), _num(t[i+5])
            tot, hon = _num(t[i+6]), _num(t[i+7])
            filas.append(_fila_unif(
                origen, prof, cuit,
                RRN=t[i],
                **{"Código PPM":       t[i+1],
                   "Ámbito/Compl.":    t[i+2],
                   "Fecha":            t[i+3],
                   "Cant.":            1,
                   "Recargo Práctica": rec,
                   "Saldo":            round(tot - ant, 2),
                   "Anticipo":         ant,
                   "Total Fact.":      tot,
                   "Honor.":           hon,
                   "Gasto":            round(tot - hon, 2),
                   "Coseguro":         ant,
                   "Forma Pago":       t[i+8],
                   "Tipo":             t[i+9]}))
            i += 10; continue
        i += 1
    return filas


def _parse_practicas_esp(html_path, prof, cuit, origen):
    """
    Detalle de Prácticas Esp. Fila: RRN, Tipo/Nro Doc, Nombre, Código(6), Cant,
    Recargo, Anticipo, Forma Pago, Coseguro, Fecha y Hora, Total Fact., Compl.,
    Saldo, Tipo Imp. (Mapeo validado fila a fila contra el ejemplo real.)
    """
    t = _tokens_html(html_path)
    filas = []; i = 0; n = len(t)
    while i < n - 13:
        if (_RRN9.match(t[i]) and _DOC.match(t[i+1]) and not _NUMTOK.match(t[i+2]) and
                _COD6.match(t[i+3]) and t[i+4].isdigit()):
            compl = "" if t[i+11] == "---" else t[i+11]
            filas.append(_fila_unif(
                origen, prof, cuit,
                RRN=t[i],
                **{"Nro/Tipo Doc":     t[i+1],
                   "Nombre Afiliado":  t[i+2],
                   "Código PPM":       t[i+3],
                   "Cant.":            _num(t[i+4]),
                   "Recargo Práctica": _num(t[i+5]),
                   "Anticipo":         _num(t[i+6]),
                   "Forma Pago":       t[i+7],
                   "Coseguro":         _num(t[i+8]),
                   "Fecha":            t[i+9],
                   "Total Fact.":      _num(t[i+10]),
                   "Ámbito/Compl.":    compl,
                   "Saldo":            _num(t[i+12]),
                   "Tipo":             t[i+13]}))
            i += 14; continue
        i += 1
    return filas


_PDF_PRAC_RE = _re.compile(
    r"^(?P<rrn>\d{9})\s+(?P<doc>\d+)\s+(?P<nombre>.+?)\s*(?P<cod>\d{6})\s+"
    r"(?P<ambito>[A-Za-z])\s+(?P<horario>\d{2}/\d{2}/\d{4}\s+\d{2}:\d{2})\s+"
    r"(?P<cant>\d+)\s+(?P<recargo>[\d.]+)\s+(?P<saldo>[\d.]+)\s+(?P<anticipo>[\d.]+)\s+"
    r"(?P<total>[\d.]+)\s+(?P<honor>[\d.]+)\s+(?P<gasto>[\d.]+)\s+"
    r"(?P<fp>\w+)\s+(?P<tipo>\w+)(?:\s+(?P<coseguro>[\d.]+))?\s*$"
)


def _lineas_pdf(pdf_path: Path) -> list:
    """Reconstruye las líneas del PDF agrupando palabras por banda vertical."""
    try:
        import pdfplumber
    except ImportError:
        log.error("pip install pdfplumber")
        return []
    from collections import defaultdict
    out = []
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            for page in pdf.pages:
                bandas = defaultdict(list)
                for w in page.extract_words():
                    bandas[round(w["top"] / 3)].append(w)
                for key in sorted(bandas):
                    ws = sorted(bandas[key], key=lambda w: w["x0"])
                    out.append(" ".join(w["text"] for w in ws))
    except Exception as e:
        log.error(f"  _lineas_pdf({pdf_path.name}): {e}")
    return out


def _parse_practicas_pdf(pdf_path, prof, cuit, origen):
    """
    Detalle de Prácticas leído del PDF del servidor (trae Nombre y Documento del
    afiliado, que el HTML no incluye). Validado contra los totales del reporte.
    """
    filas = []
    for ln in _lineas_pdf(Path(pdf_path)):
        m = _PDF_PRAC_RE.match(ln.strip())
        if not m:
            continue
        d = m.groupdict()
        filas.append(_fila_unif(
            origen, prof, cuit,
            RRN=d["rrn"],
            **{"Nro/Tipo Doc":     d["doc"],
               "Nombre Afiliado":  d["nombre"].strip(),
               "Código PPM":       d["cod"],
               "Ámbito/Compl.":    d["ambito"],
               "Fecha":            d["horario"],
               "Cant.":            _num(d["cant"]),
               "Recargo Práctica": _num(d["recargo"]),
               "Saldo":            _num(d["saldo"]),
               "Anticipo":         _num(d["anticipo"]),
               "Total Fact.":      _num(d["total"]),
               "Honor.":           _num(d["honor"]),
               "Gasto":            _num(d["gasto"]),
               "Coseguro":         _num(d["coseguro"] or 0),
               "Forma Pago":       d["fp"],
               "Tipo":             d["tipo"]}))
    return filas


def parsear_por_origen(archivo: Path, prof: str, cuit: str, origen: str) -> list:
    """
    Despacha al parser correcto según el origen y el tipo de archivo, y devuelve
    filas unificadas. Prácticas se lee del PDF; el resto del HTML.
    """
    o = (origen or "").lower()
    archivo = Path(archivo)
    try:
        if str(archivo).lower().endswith(".pdf"):
            # Hoy solo Prácticas se baja en PDF.
            return _parse_practicas_pdf(archivo, prof, cuit, origen or "Prácticas")
        if "esp" in o:
            return _parse_practicas_esp(archivo, prof, cuit, origen)
        if "práctica" in o or "practica" in o:
            return _parse_practicas(archivo, prof, cuit, origen)
        return _parse_ambulatorio(archivo, prof, cuit, origen or "Ambulatorio")
    except Exception as e:
        log.error(f"  parsear_por_origen[{origen}] {archivo.name}: {e}")
        return []


def consolidar_excel(todos_resultados: list, output_dir: Path) -> Path | None:
    """
    Genera UN Excel unificado con los tres orígenes a nivel detalle, en columnas
    normalizadas (lo que falta queda en blanco/0). Hojas:
      • Detalle      → todas las prestaciones, con columna Origen
      • Totalizador  → suma del detalle por Origen + Código PPM
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        log.error("openpyxl no instalado.")
        return None

    log.info("\nGenerando Excel consolidado (unificado por origen)...")

    filas = []
    for r in todos_resultados:
        if r.get("estado") != "ok" or not r.get("html"):
            continue
        html_path = Path(r["html"])
        if not html_path.exists():
            continue
        prof   = r.get("profesional", "")
        cuit   = str(r.get("cuit", ""))
        origen = r.get("origen", "Ambulatorio")
        filas.extend(parsear_por_origen(html_path, prof, cuit, origen))

    if not filas:
        log.warning("  Sin datos para consolidar (HTMLs vacíos o no parseables).")
        return None

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill    = PatternFill("solid", fgColor="D6E4F0")
    center      = Alignment(horizontal="center", vertical="center")

    def estilizar(ws, headers):
        ws.row_dimensions[1].height = 20
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = center
        ws.freeze_panes = "A2"

    def autoajustar(ws):
        for col in ws.columns:
            m = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(m + 2, 42)

    wb = openpyxl.Workbook()

    # ── Hoja Detalle (unificada) ────────────────────────────────────────────
    ws = wb.active; ws.title = "Detalle"
    estilizar(ws, COLS_UNIF)
    for ri, fila in enumerate(filas, 2):
        fill = alt_fill if ri % 2 == 0 else None
        for ci, key in enumerate(COLS_UNIF, 1):
            cell = ws.cell(row=ri, column=ci, value=fila.get(key, ""))
            if fill:
                cell.fill = fill
    autoajustar(ws)

    # ── Hoja Totalizador (suma del detalle por Profesional + Origen + Código) ─
    from collections import OrderedDict
    NUMER = ["Cant.", "Recargo Práctica", "Saldo", "Anticipo", "Total Fact.", "Honor.", "Gasto", "Coseguro"]
    agg = OrderedDict()
    for f in filas:
        k = (f.get("Profesional", ""), str(f.get("CUIT", "")),
             f.get("Origen", ""), str(f.get("Código PPM", "")), f.get("Descripción PPM", ""))
        a = agg.setdefault(k, {c: 0.0 for c in NUMER})
        for c in NUMER:
            try: a[c] += float(f.get(c) or 0)
            except Exception: pass

    headers_tot = ["Profesional", "CUIT", "Origen", "Código PPM", "Descripción PPM"] + NUMER
    ws2 = wb.create_sheet("Totalizador")
    estilizar(ws2, headers_tot)
    for ri, (k, a) in enumerate(agg.items(), 2):
        fill = alt_fill if ri % 2 == 0 else None
        valores = [k[0], k[1], k[2], k[3], k[4]] + [round(a[c], 2) for c in NUMER]
        for ci, val in enumerate(valores, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill
    autoajustar(ws2)

    ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta_xlsx = output_dir / f"Consolidado_{MES}_{ANIO}_{ts}.xlsx"
    wb.save(str(ruta_xlsx))
    por_origen = {}
    for f in filas:
        por_origen[f["Origen"]] = por_origen.get(f["Origen"], 0) + 1
    log.info(f"  ✓ Excel consolidado: {ruta_xlsx}")
    log.info(f"  Filas: {len(filas)} | por origen: {por_origen} | códigos: {len(agg)}")
    return ruta_xlsx


async def main():
    log.info(f"{'='*56}\nBot Conexia SEROS — {MES} {ANIO}\n{'='*56}")
    grupos = leer_excel(EXCEL, DESDE, HASTA)
    todos  = []

    async with async_playwright() as pw:
        browser = await pw.chromium.launch(
            headless=HEADLESS,
            args=[
                "--no-sandbox",
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--disable-gpu",
                "--memory-pressure-off",
                "--disable-background-networking",
                "--disable-extensions",
                "--no-first-run",
            ])
        ctx = await browser.new_context(
            viewport={"width":1280,"height":800}, locale="es-AR",
            timezone_id="America/Argentina/Buenos_Aires",
            user_agent=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                        "AppleWebKit/537.36 (KHTML, like Gecko) "
                        "Chrome/124.0.0.0 Safari/537.36"),
            accept_downloads=True, ignore_https_errors=True)

        lista = list(grupos.items())
        for i, ((u, c), prests) in enumerate(lista, 1):
            log.info(f"\n[{i}/{len(lista)}] {u}")
            try:
                # Timeout por usuario: 3 min base + 2 min por registro
                _timeout = min(180 + len(prests) * 120, 1800)  # máx 30 min
                r = await asyncio.wait_for(
                    procesar_usuario(ctx, u, c, prests),
                    timeout=_timeout
                )
                todos.extend(r)
            except asyncio.TimeoutError:
                log.error(f"  TIMEOUT — {u} cancelado automáticamente (timeout={_timeout}s)")
                for p in prests:
                    todos.append({"usuario":u,"profesional":p["nombre"],"cuit":p["cuit"],
                                  "html":None,"pdf":None,"estado":"error_timeout",
                                  "detalle":"Superó 8 minutos"})
                # Cerrar todas las páginas abiertas y crear una nueva limpia
                try:
                    for pg in ctx.pages:
                        await pg.close()
                except Exception:
                    pass
            except Exception as e:
                log.error(f"Error {u}: {e}")
                for p in prests:
                    todos.append({"usuario":u,"profesional":p["nombre"],"cuit":p["cuit"],
                                  "html":None,"pdf":None,"estado":"error_inesperado","detalle":str(e)})
                # También limpiar páginas en caso de error grave
                try:
                    for pg in ctx.pages:
                        await pg.close()
                except Exception:
                    pass
            if i < len(lista):
                await asyncio.sleep(random.uniform(8, 15))

        await browser.close()

        # ── HTML → PDF ────────────────────────────────────────────────────────
        htmls = list(SALIDA.rglob("*.html"))
        if htmls:
            log.info(f"\nCONVIRTIENDO {len(htmls)} HTML → PDF")
            async with async_playwright() as pw2:
                for k, h in enumerate(htmls, 1):
                    log.info(f"  [{k}/{len(htmls)}] {h.name}")
                    pdf = await html_a_pdf(h, pw2)
                    for r in todos:
                        if r.get("html") and Path(r["html"]) == h:
                            r["pdf"] = str(pdf) if pdf else None; break

    fall = guardar_reporte(todos)
    ok   = sum(1 for r in todos if r.get("estado") == "ok")
    log.info(f"\nRESUMEN: {len(todos)} | OK:{ok} | Fallidos:{len(fall)}")

    # ── ZIP de PDFs ───────────────────────────────────────────────────────────
    import zipfile as _zf
    pdfs = list(SALIDA.rglob("*.pdf"))
    if pdfs:
        zip_path = SALIDA / f"PDFs_{MES}_{ANIO}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        with _zf.ZipFile(zip_path, "w", _zf.ZIP_DEFLATED) as zf:
            for p in pdfs:
                zf.write(p, p.name)
        log.info(f"ZIP PDFs: {zip_path.name} ({len(pdfs)} archivos)")

    # ── Excel consolidado ─────────────────────────────────────────────────────
    # Usar la función interna consolidar_excel (parsea los HTML descargados y
    # arma el Excel). Antes se intentaba importar un módulo externo
    # 'generar_excel' que, al no existir, hacía que no se consolidara nada.
    try:
        ruta_xlsx = consolidar_excel(todos, SALIDA)
        if ruta_xlsx is None:
            log.warning("No se generó el Excel consolidado (sin datos parseables).")
    except Exception as e:
        log.error(f"Error al generar Excel consolidado: {e}")
        import traceback; log.debug(traceback.format_exc())

    log.info(f"Salida: {SALIDA.resolve()}")
if __name__=="__main__": asyncio.run(main())
