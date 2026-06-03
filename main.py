"""
main.py — API FastAPI para Bot Conexia SEROS
Corre en PikaPods, expone endpoints para el frontend en Vercel.
"""
import os, subprocess, json, re, sys, threading, uuid, logging
from pathlib import Path
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(message)s")
log = logging.getLogger("conexia_api")

app = FastAPI(title="Conexia SEROS API")

# ── CORS: permitir llamadas desde cualquier origen ───────────────────────────
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── Rutas ─────────────────────────────────────────────────────────────────────
UPLOAD_DIR = Path("/app/uploads")
OUTPUT_DIR = Path("/app/output")
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
EXCEL_PATH = UPLOAD_DIR / "credenciales_seros.xlsx"

# ── Estado global de ejecución ────────────────────────────────────────────────
# Estado por parte (1, 2, 3)
def _estado_inicial():
    return {
        "status":   "idle",   # idle | running | done | error
        "inicio":   None,
        "fin":      None,
        "log":      [],
        "ok":       0,
        "fallidos": 0,
        "carpeta":  None,
        "parte":    0,
        "desde":    0,
        "hasta":    0,
    }

estados = {1: _estado_inicial(), 2: _estado_inicial(), 3: _estado_inicial(), 4: _estado_inicial()}

# Cuántos registros tiene el Excel actual
_total_registros = 0


def _contar_registros():
    """Cuenta los registros del Excel subido."""
    global _total_registros
    if not EXCEL_PATH.exists():
        return 0
    try:
        import openpyxl as _opx
        wb = _opx.load_workbook(EXCEL_PATH, read_only=True)
        ws = wb.active
        _total_registros = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if any(c for c in r))
        return _total_registros
    except Exception:
        return 0


def _rangos_partes(total: int):
    """Divide total registros en 4 partes lo más iguales posible."""
    n = 4
    base = total // n
    resto = total % n
    partes = []
    inicio = 1
    for i in range(n):
        extra = 1 if i < resto else 0
        fin = inicio + base + extra - 1
        partes.append((inicio, min(fin, total)))
        inicio = fin + 1
    return partes  # [(1,30),(31,60),(61,90),(91,119)]


def _run_parte(parte: int, desde: int, hasta: int):
    """Corre el bot para un rango de registros."""
    e = estados[parte]
    e["status"]   = "running"
    e["inicio"]   = datetime.now().isoformat()
    e["log"]      = []
    e["ok"]       = 0
    e["fallidos"] = 0
    e["parte"]    = parte
    e["desde"]    = desde
    e["hasta"]    = hasta

    # Limpiar procesos zombie y liberar recursos antes de iniciar
    import gc, time
    gc.collect()
    time.sleep(3)  # pausa para que el SO libere recursos del proceso anterior

    env = os.environ.copy()
    env["EXCEL_PATH"] = str(EXCEL_PATH)
    env["OUTPUT_DIR"] = str(OUTPUT_DIR)
    env["HEADLESS"]   = "true"
    env["DESDE"]      = str(desde)
    env["HASTA"]      = str(hasta)
    # Limitar memoria de Chromium
    env["PLAYWRIGHT_CHROMIUM_ARGS"] = "--disable-dev-shm-usage --no-sandbox --memory-pressure-off"

    try:
        proc = subprocess.Popen(
            [sys.executable, "conexia_bot.py"],
            stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace", env=env, bufsize=1
        )
        for line in proc.stdout:
            line = line.rstrip()
            e["log"].append(line)
            if len(e["log"]) > 500:
                e["log"] = e["log"][-500:]
            if "✓" in line and ".html" in line:
                e["ok"] += 1
            if "Credenciales inválidas" in line or "TIMEOUT" in line:
                e["fallidos"] += 1
        proc.wait()
        e["status"] = "done" if proc.returncode == 0 else "error"
    except Exception as ex:
        e["log"].append(f"ERROR: {ex}")
        e["status"] = "error"

    e["fin"] = datetime.now().isoformat()
    # Guardar carpeta de salida más reciente
    subcarpetas = sorted(
        [d for d in OUTPUT_DIR.iterdir() if d.is_dir() and d.name[:8].isdigit()],
        reverse=True
    ) if OUTPUT_DIR.exists() else []
    e["carpeta"] = str(subcarpetas[0]) if subcarpetas else None

    # Re-unificar automáticamente con TODAS las partes completadas.
    # Así el Excel unificado siempre refleja todas las partes 'done' sin
    # depender de que el frontend llame a /unificar en el momento justo
    # (antes solo quedaba consolidada la parte 1, la única 'done' cuando
    # el frontend llamaba a /unificar).
    if e["status"] == "done":
        try:
            _unificar_background()
        except Exception as ex:
            log.error(f"Error al re-unificar tras la parte {parte}: {ex}")


# ── Endpoints ─────────────────────────────────────────────────────────────────

import hashlib, secrets
from pydantic import BaseModel


def _merge_consolidados(partes_done: list, carpeta_unif: Path, ts: str) -> Path | None:
    """
    Une los Consolidado_*.xlsx que generó cada parte (hojas Detalle y
    Totalizador) en un único Excel unificado. No depende de módulos externos.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    center      = Alignment(horizontal="center", vertical="center")

    detalle_rows, total_rows = [], []
    headers_det, headers_tot = None, None

    for p in partes_done:
        carpeta = Path(estados[p]["carpeta"])
        xlsxs = sorted(carpeta.glob("Consolidado_*.xlsx"), reverse=True)
        if not xlsxs:
            log.warning(f"Parte {p}: sin Consolidado_*.xlsx en {carpeta}")
            continue
        try:
            wb = openpyxl.load_workbook(xlsxs[0], read_only=True, data_only=True)
        except Exception as ex:
            log.error(f"Parte {p}: no se pudo leer {xlsxs[0].name}: {ex}")
            continue
        if "Detalle" in wb.sheetnames:
            filas = list(wb["Detalle"].iter_rows(values_only=True))
            if filas:
                if headers_det is None:
                    headers_det = filas[0]
                detalle_rows.extend(filas[1:])
        if "Totalizador" in wb.sheetnames:
            filas = list(wb["Totalizador"].iter_rows(values_only=True))
            if filas:
                if headers_tot is None:
                    headers_tot = filas[0]
                total_rows.extend(filas[1:])
        wb.close()

    if not detalle_rows and not total_rows:
        return None

    def _escribir(ws, headers, filas):
        if headers:
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=1, column=ci, value=h)
                c.font = header_font; c.fill = header_fill; c.alignment = center
            ws.freeze_panes = "A2"
        for ri, fila in enumerate(filas, 2):
            for ci, val in enumerate(fila, 1):
                ws.cell(row=ri, column=ci, value=val)

    out = openpyxl.Workbook()
    ws_det = out.active; ws_det.title = "Detalle"
    _escribir(ws_det, headers_det, detalle_rows)
    ws_tot = out.create_sheet("Totalizador")
    _escribir(ws_tot, headers_tot, total_rows)

    ruta = carpeta_unif / f"Consolidado_unificado_{ts}.xlsx"
    out.save(str(ruta))
    log.info(f"Excel unificado: {ruta.name} "
             f"({len(detalle_rows)} detalle, {len(total_rows)} totalizador)")
    return ruta



# ── Usuarios: configurar via variables de entorno en Railway ──────────────────
# Formato: USERS=usuario1:clave1,usuario2:clave2
def _cargar_usuarios():
    raw = os.getenv("USERS", "admin:seros2026")
    usuarios = {}
    for par in raw.split(","):
        if ":" in par:
            u, c = par.strip().split(":", 1)
            usuarios[u.strip()] = c.strip()
    return usuarios

# Tokens de sesión activos (en memoria)
_tokens_validos: set[str] = set()

class LoginData(BaseModel):
    usuario: str
    clave:   str

@app.get("/")
def root():
    return {"app": "Conexia SEROS API", "version": "1.0"}

@app.post("/login")
def login(data: LoginData):
    usuarios = _cargar_usuarios()
    if data.usuario in usuarios and usuarios[data.usuario] == data.clave:
        token = secrets.token_hex(32)
        _tokens_validos.add(token)
        return {"ok": True, "token": token}
    raise HTTPException(401, "Credenciales incorrectas")


def _unificar_background():
    """Unifica los resultados de las 4 partes en background."""
    import zipfile, shutil
    try:
        partes_done = [p for p, e in estados.items() if e["status"] == "done" and e["carpeta"]]
        if not partes_done:
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        carpeta_unif = OUTPUT_DIR / f"unificado_{ts}"
        carpeta_unif.mkdir(parents=True, exist_ok=True)

        # Reunir HTMLs
        for p in partes_done:
            for h in Path(estados[p]["carpeta"]).rglob("*.html"):
                dest = carpeta_unif / h.name
                if dest.exists():
                    dest = carpeta_unif / f"{h.stem}_p{p}{h.suffix}"
                shutil.copy2(h, dest)

        # ZIP de PDFs
        todos_pdfs = []
        for p in partes_done:
            todos_pdfs.extend(Path(estados[p]["carpeta"]).rglob("*.pdf"))
        zip_path = carpeta_unif / f"PDFs_Unificado_{ts}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            for pdf in todos_pdfs:
                zf.write(pdf, pdf.name)

        # Excel consolidado uniendo los Consolidado_*.xlsx de cada parte
        try:
            _merge_consolidados(partes_done, carpeta_unif, ts)
        except Exception as ex:
            log.error(f"Error generando Excel unificado: {ex}")

        log.info(f"✓ Unificación completada con partes {partes_done}: {carpeta_unif}")

        # Limpiar carpetas 'unificado_*' anteriores para no acumular disco;
        # se conserva solo la recién creada (la más completa).
        try:
            for d in OUTPUT_DIR.glob("unificado_*"):
                if d.is_dir() and d != carpeta_unif:
                    shutil.rmtree(d, ignore_errors=True)
        except Exception as ex:
            log.warning(f"No se pudieron limpiar unificados viejos: {ex}")
    except Exception as e:
        log.error(f"Error en unificación background: {e}")


@app.post("/reset")
def reset_estado():
    """Reinicia el estado de las 3 partes para una nueva ejecución."""
    global _total_registros
    for p in [1, 2, 3, 4]:
        estados[p] = _estado_inicial()
    _total_registros = 0
    return {"ok": True}


@app.post("/upload")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Solo se aceptan archivos .xlsx")
    contenido = await file.read()
    EXCEL_PATH.write_bytes(contenido)
    total = _contar_registros()
    partes = _rangos_partes(total) if total > 0 else [(1,40),(41,80),(81,119)]
    return {
        "ok": True,
        "registros": total,
        "partes": [
            {"parte": i+1, "desde": p[0], "hasta": p[1], "cantidad": p[1]-p[0]+1}
            for i, p in enumerate(partes)
        ]
    }


@app.post("/run/{parte}")
def run_parte(parte: int):
    """Inicia la extracción de una parte (1, 2 o 3)."""
    if parte not in [1, 2, 3, 4]:
        raise HTTPException(400, "Parte debe ser 1, 2, 3 o 4")
    e = estados[parte]
    if e["status"] == "running":
        raise HTTPException(409, f"La parte {parte} ya está ejecutándose")
    # Verificar que no haya otra parte corriendo
    for p, st in estados.items():
        if p != parte and st["status"] == "running":
            raise HTTPException(409, f"La parte {p} está ejecutándose. Esperá que termine.")
    if not EXCEL_PATH.exists():
        raise HTTPException(400, "Primero subí el Excel de credenciales")

    total = _contar_registros()
    partes = _rangos_partes(total)
    desde, hasta = partes[parte - 1]

    threading.Thread(target=_run_parte, args=(parte, desde, hasta), daemon=True).start()
    return {"ok": True, "parte": parte, "desde": desde, "hasta": hasta}


@app.get("/status")
def get_status():
    """Estado de las 3 partes."""
    partes_status = {}
    for p, e in estados.items():
        partes_status[str(p)] = {
            "status":   e["status"],
            "inicio":   e["inicio"],
            "fin":      e["fin"],
            "ok":       e["ok"],
            "fallidos": e["fallidos"],
            "desde":    e["desde"],
            "hasta":    e["hasta"],
            "log_tail": e["log"][-30:],
        }
    total = _contar_registros()
    partes_info = _rangos_partes(total) if total > 0 else [(1,40),(41,80),(81,119)]
    return {
        "excel_ok":    EXCEL_PATH.exists(),
        "total":       total,
        "partes_info": [{"parte": i+1, "desde": p[0], "hasta": p[1]} for i, p in enumerate(partes_info)],
        "partes":      partes_status,
    }


@app.get("/log")
def get_log():
    """Log completo de la última ejecución, por parte."""
    return {"log": {str(p): estados[p]["log"] for p in estados}}


def _buscar_en_carpeta(carpeta_str: str | None, patron: str) -> Path | None:
    """Busca un archivo por patrón. Si no hay carpeta en memoria, busca en todo output/."""
    candidatos = []
    if carpeta_str:
        carpeta = Path(carpeta_str)
        candidatos = sorted(carpeta.glob(patron), reverse=True)
    # Fallback: buscar en todas las subcarpetas de output
    if not candidatos and OUTPUT_DIR.exists():
        candidatos = sorted(OUTPUT_DIR.rglob(patron), reverse=True)
    return candidatos[0] if candidatos else None


@app.get("/download/unificado/excel")
def download_unif_excel_alias():
    return download_unif_excel()

@app.get("/download/unificado/pdfs")
def download_unif_pdfs_alias():
    return download_unif_pdfs()

@app.get("/download/unificado/fallidos")
def download_unif_fallidos():
    """Une los archivos FALLIDOS de todas las partes en uno solo."""
    lineas = []
    for p in [1, 2, 3, 4]:
        f = _buscar_en_carpeta(estados[p]["carpeta"], "FALLIDOS_*.txt")
        if f:
            lineas.append(f"=== PARTE {p} (registros {estados[p]['desde']}-{estados[p]['hasta']}) ===")
            lineas.append(f.read_text(encoding="utf-8"))
            lineas.append("")
    if not lineas:
        todos = sorted(OUTPUT_DIR.rglob("FALLIDOS_*.txt"), reverse=True)
        for f in todos:
            lineas.append(f"=== {f.parent.name} ===")
            lineas.append(f.read_text(encoding="utf-8"))
            lineas.append("")
    if not lineas:
        raise HTTPException(404, "No hay archivos de fallidos")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = OUTPUT_DIR / f"FALLIDOS_consolidado_{ts}.txt"
    ruta.write_text("\n".join(lineas), encoding="utf-8")
    return FileResponse(str(ruta), filename=ruta.name, media_type="text/plain")

@app.get("/download/{parte}/excel")
def download_excel(parte: int):
    if parte not in [1,2,3,4]:
        raise HTTPException(400, "Parte inválida")
    f = _buscar_en_carpeta(estados[parte]["carpeta"], "Consolidado_*.xlsx")
    if not f:
        raise HTTPException(404, f"Excel parte {parte} no disponible")
    return FileResponse(str(f), filename=f.name,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/download/{parte}/pdfs")
def download_pdfs(parte: int):
    if parte not in [1,2,3,4]:
        raise HTTPException(400, "Parte inválida")
    f = _buscar_en_carpeta(estados[parte]["carpeta"], "PDFs_*.zip")
    if not f:
        raise HTTPException(404, f"ZIP parte {parte} no disponible")
    return FileResponse(str(f), filename=f.name, media_type="application/zip")


@app.get("/download/{parte}/fallidos")
def download_fallidos(parte: int):
    if parte not in [1,2,3,4]:
        raise HTTPException(400, "Parte inválida")
    f = _buscar_en_carpeta(estados[parte]["carpeta"], "FALLIDOS_*.txt")
    if not f:
        raise HTTPException(404, f"Fallidos parte {parte} no disponible")
    return FileResponse(str(f), filename=f.name, media_type="text/plain")


@app.post("/unificar")
def unificar():
    """Une los resultados de las 3 partes en un Excel y ZIP consolidado."""
    import zipfile, shutil

    # Verificar que al menos una parte terminó
    partes_done = [p for p, e in estados.items() if e["status"] == "done" and e["carpeta"]]
    if not partes_done:
        raise HTTPException(400, "Ninguna parte completada aún")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    carpeta_unif = OUTPUT_DIR / f"unificado_{ts}"
    carpeta_unif.mkdir(parents=True, exist_ok=True)

    # Reunir todos los HTMLs de las partes completadas
    todos_htmls = []
    for p in partes_done:
        carpeta = Path(estados[p]["carpeta"])
        htmls = list(carpeta.rglob("*.html"))
        for h in htmls:
            dest = carpeta_unif / h.name
            # Evitar colisiones de nombre
            if dest.exists():
                dest = carpeta_unif / f"{h.stem}_p{p}{h.suffix}"
            shutil.copy2(h, dest)
            todos_htmls.append(dest)

    # ZIP unificado de PDFs
    todos_pdfs = []
    for p in partes_done:
        carpeta = Path(estados[p]["carpeta"])
        todos_pdfs.extend(carpeta.rglob("*.pdf"))

    zip_path = carpeta_unif / f"PDFs_Unificado_{ts}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for pdf in todos_pdfs:
            zf.write(pdf, pdf.name)

    # Generar Excel consolidado uniendo los Consolidado_*.xlsx de cada parte
    try:
        ruta_excel = _merge_consolidados(partes_done, carpeta_unif, ts)
        excel_ok = ruta_excel is not None
    except Exception as ex:
        log.error(f"Error al unir Excel consolidado: {ex}")
        excel_ok = False

    return {
        "ok": True,
        "partes_unificadas": partes_done,
        "pdfs": len(todos_pdfs),
        "htmls": len(todos_htmls),
        "excel_ok": excel_ok,
        "carpeta": str(carpeta_unif)
    }


@app.get("/download/unificado/excel_full")
def download_unif_excel():
    carpetas = sorted(
        [d for d in OUTPUT_DIR.iterdir() if d.is_dir() and d.name.startswith("unificado_")],
        reverse=True
    ) if OUTPUT_DIR.exists() else []
    if not carpetas:
        raise HTTPException(404, "No hay resultado unificado aún")
    f = next(iter(sorted(carpetas[0].glob("Consolidado_*.xlsx"), reverse=True)), None)
    if not f:
        raise HTTPException(404, "Excel unificado no disponible")
    return FileResponse(str(f), filename=f.name,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/download/unificado/pdfs_full")
def download_unif_pdfs():
    carpetas = sorted(
        [d for d in OUTPUT_DIR.iterdir() if d.is_dir() and d.name.startswith("unificado_")],
        reverse=True
    ) if OUTPUT_DIR.exists() else []
    if not carpetas:
        raise HTTPException(404, "No hay resultado unificado aún")
    f = next(iter(sorted(carpetas[0].glob("PDFs_Unificado_*.zip"), reverse=True)), None)
    if not f:
        raise HTTPException(404, "ZIP unificado no disponible")
    return FileResponse(str(f), filename=f.name, media_type="application/zip")


# ══════════════════════════════════════════════════════════════════════════════
# BOT 2 — EXTRANET ISSyS (Control de facturas y pagos)
# ══════════════════════════════════════════════════════════════════════════════

estado_extranet = {
    "status":   "idle",
    "inicio":   None,
    "fin":      None,
    "log":      [],
    "total":    0,
    "procesados": 0,
    "datos":    [],          # resultados en memoria
}

# Archivo de comprobantes ARCA (persistencia simple en JSON)
COMP_PATH = Path("/app/uploads/comprobantes_arca.json")

def _cargar_comprobantes() -> dict:
    """Carga el mapa {clave_unica: n_comprobante} desde disco."""
    if COMP_PATH.exists():
        try:
            return json.loads(COMP_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def _guardar_comprobantes(comp: dict):
    COMP_PATH.write_text(json.dumps(comp, ensure_ascii=False, indent=2), encoding="utf-8")


def _run_extranet():
    """Corre el bot de extranet en thread separado."""
    import subprocess, sys
    estado_extranet["status"]    = "running"
    estado_extranet["inicio"]    = datetime.now().isoformat()
    estado_extranet["log"]       = []
    estado_extranet["procesados"]= 0

    env = os.environ.copy()
    env["HEADLESS"] = "true"

    try:
        proc = subprocess.Popen(
            [sys.executable, "extranet_bot.py"],
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True, encoding="utf-8", errors="replace",
            env=env, bufsize=1
        )
        for line in proc.stdout:
            line = line.rstrip()
            estado_extranet["log"].append(line)
            if len(estado_extranet["log"]) > 500:
                estado_extranet["log"] = estado_extranet["log"][-500:]
            if "Procesando" in line:
                m = re.search(r"Procesando (\d+)", line)
                if m:
                    estado_extranet["total"] = int(m.group(1))
            if "Extrayendo:" in line:
                estado_extranet["procesados"] += 1

        proc.wait()
        estado_extranet["status"] = "done" if proc.returncode == 0 else "error"

        # Cargar el JSON generado
        jsons = sorted(OUTPUT_DIR.rglob("extranet_*.json"), reverse=True)
        if jsons:
            datos = json.loads(jsons[0].read_text(encoding="utf-8"))
            # Inyectar comprobantes guardados
            comp = _cargar_comprobantes()
            for miembro in datos:
                for det in miembro.get("detalle", []):
                    clave = f"{miembro['cuit']}_{det.get('nro_pedido_actuacion','')}_{det.get('nro','')}"
                    det["n_comprobante_arca"] = comp.get(clave, "")
            estado_extranet["datos"] = datos

    except Exception as e:
        estado_extranet["log"].append(f"ERROR: {e}")
        estado_extranet["status"] = "error"

    estado_extranet["fin"] = datetime.now().isoformat()


@app.post("/extranet/upload-json")
async def extranet_upload_json(file: UploadFile = File(...)):
    """Sube el JSON generado por extranet_bot.py al servidor."""
    if not file.filename.endswith(".json"):
        raise HTTPException(400, "Solo se aceptan archivos .json")
    contenido = await file.read()
    # Guardar en output/
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = OUTPUT_DIR / f"extranet_{ts}.json"
    ruta.write_bytes(contenido)
    # Cargar en memoria
    datos = json.loads(contenido.decode("utf-8"))
    comp  = _cargar_comprobantes()
    for miembro in datos:
        for det in miembro.get("detalle", []):
            clave = f"{miembro['cuit']}_{det.get('nro_pedido_actuacion','')}_{det.get('nro','')}"
            det["n_comprobante_arca"] = comp.get(clave, "")
    estado_extranet["datos"] = datos
    return {"ok": True, "miembros": len(datos), "archivo": ruta.name}


@app.post("/extranet/run")
def extranet_run():
    if estado_extranet["status"] == "running":
        raise HTTPException(409, "Ya hay una extracción en curso")
    if not os.getenv("EXTRANET_USER") or not os.getenv("EXTRANET_PASS"):
        raise HTTPException(400, "Faltan EXTRANET_USER y EXTRANET_PASS en variables de entorno")
    threading.Thread(target=_run_extranet, daemon=True).start()
    return {"ok": True}


@app.get("/extranet/status")
def extranet_status():
    return {
        "status":     estado_extranet["status"],
        "inicio":     estado_extranet["inicio"],
        "fin":        estado_extranet["fin"],
        "total":      estado_extranet["total"],
        "procesados": estado_extranet["procesados"],
        "log_tail":   estado_extranet["log"][-50:],
    }


@app.get("/extranet/datos")
def extranet_datos(profesional: str = "", periodo: str = ""):
    """Devuelve los datos filtrados por profesional y/o periodo."""
    datos = estado_extranet["datos"]
    if not datos:
        # Intentar cargar el JSON más reciente
        jsons = sorted(OUTPUT_DIR.rglob("extranet_*.json"), reverse=True)
        if jsons:
            datos = json.loads(jsons[0].read_text(encoding="utf-8"))
            comp  = _cargar_comprobantes()
            for miembro in datos:
                for det in miembro.get("detalle", []):
                    clave = f"{miembro['cuit']}_{det.get('nro_pedido_actuacion','')}_{det.get('nro','')}"
                    det["n_comprobante_arca"] = comp.get(clave, "")
            estado_extranet["datos"] = datos

    resultado = datos
    if profesional:
        resultado = [m for m in resultado
                     if profesional.lower() in m.get("nombre","").lower()]
    if periodo:
        resultado_filtrado = []
        for m in resultado:
            m_copy = dict(m)
            m_copy["detalle"] = [
                d for d in m.get("detalle", [])
                if periodo in str(d.get("fecha", "")) or
                   periodo in str(d.get("fecha_gasto", ""))
            ]
            if m_copy["detalle"] or not m.get("detalle"):
                resultado_filtrado.append(m_copy)
        resultado = resultado_filtrado

    return {"datos": resultado, "total": len(resultado)}


class ComprobantePatch(BaseModel):
    cuit:                str
    nro_pedido_actuacion: str
    nro:                 str
    n_comprobante_arca:  str


@app.post("/extranet/comprobante")
def guardar_comprobante(patch: ComprobantePatch):
    """Guarda el N° de comprobante ARCA para una presentación específica."""
    comp  = _cargar_comprobantes()
    clave = f"{patch.cuit}_{patch.nro_pedido_actuacion}_{patch.nro}"
    comp[clave] = patch.n_comprobante_arca

    # Actualizar en memoria también
    for miembro in estado_extranet["datos"]:
        if miembro.get("cuit") == patch.cuit:
            for det in miembro.get("detalle", []):
                if (det.get("nro_pedido_actuacion") == patch.nro_pedido_actuacion and
                        det.get("nro") == patch.nro):
                    det["n_comprobante_arca"] = patch.n_comprobante_arca

    _guardar_comprobantes(comp)
    return {"ok": True, "clave": clave}
