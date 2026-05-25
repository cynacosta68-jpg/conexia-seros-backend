"""
generar_excel.py — Parsea los HTML descargados y genera el Excel consolidado.
Ejecutar desde la carpeta conexia_bot:
    python generar_excel.py
"""
import re, logging
from pathlib import Path
from datetime import date
from dateutil.relativedelta import relativedelta

# ── Logger ────────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)-8s %(message)s")
log = logging.getLogger("consolidar")

# ── Período ───────────────────────────────────────────────────────────────────
_ant  = date.today() - relativedelta(months=1)
MESES = {1:"Enero",2:"Febrero",3:"Marzo",4:"Abril",5:"Mayo",6:"Junio",
         7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
MES  = MESES[_ant.month]
ANIO = str(_ant.year)

_base = Path("output")
# Usar la subcarpeta más reciente (formato YYYYMMDD_HHmm) si existe
# o "output" directamente si se llama de forma standalone
def _get_output():
    subcarpetas = sorted([d for d in _base.iterdir()
                          if d.is_dir() and d.name[:8].isdigit()], reverse=True)
    return subcarpetas[0] if subcarpetas else _base

OUTPUT = _get_output()

# ── Regex ─────────────────────────────────────────────────────────────────────
AFIL_RE   = re.compile(r"^\d{8,10}$")
FECHA_RE  = re.compile(r"^\d{2}/\d{2}/\d{4}$")
CODIGO_RE = re.compile(r"^[A-Z]\d{7}$")
NUM_RE    = re.compile(r"^\d+\.?\d*$")
WEBP_RE   = re.compile(r"^WEBP\d+$")


def parsear_html(html_path: Path, nombre: str, cuit: str) -> dict:
    from bs4 import BeautifulSoup

    contenido = html_path.read_bytes()
    texto = ""
    for enc in ("utf-8", "latin-1", "cp1252"):
        try:
            texto = contenido.decode(enc); break
        except UnicodeDecodeError:
            pass
    if not texto:
        log.warning(f"  No se pudo decodificar {html_path.name}")
        return {"detalle": [], "totalizador": []}

    soup = BeautifulSoup(texto, "lxml")
    tablas = soup.find_all("table")
    if not tablas:
        return {"detalle": [], "totalizador": []}

    log.info(f"  {html_path.name}: {len(tablas)} tablas")

    # Extraer todos los <td> no vacíos de la tabla principal
    tokens = []
    for td in tablas[0].find_all("td"):
        v = td.get_text(" ", strip=True).replace("\xa0", " ").strip()
        if v:
            tokens.append(v)

    detalle = []
    totalizador = []
    i = 0
    n = len(tokens)

    while i < n:
        t = tokens[i]

        # ── FILA DE DETALLE ───────────────────────────────────────────────────
        # Estructura: [afil][nombre][cod_ppm][desc][fecha][rrn][manual][cant]
        #             [anticipo][saldo][mutual][total][imp_gastos][imp_honor][terminal]
        if (AFIL_RE.match(t) and
                i + 4 < n and
                FECHA_RE.match(tokens[i + 4]) and
                CODIGO_RE.match(tokens[i + 2])):

            def g(o, d=""): return tokens[i + o] if i + o < n else d
            def gn(o):
                try:    return float(g(o).replace(",", "."))
                except: return 0.0

            detalle.append({
                "N° Afiliado":     g(0),
                "Nombre Afiliado": g(1),
                "Código PPM":      g(2),
                "Descripción PPM": g(3),
                "Fecha":           g(4),
                "RRN":             g(5),
                "Manual":          g(6),
                "Nro Orden":       "",
                "Cant":            gn(7),
                "Anticipo":        gn(8),
                "Saldo":           gn(9),
                "Mutual":          gn(10),
                "Total Facturado": gn(11),
                "Imp Gastos":      gn(12),
                "Imp Honor":       gn(13),
                "Terminal":        g(14) if WEBP_RE.match(g(14)) else "",
                "Profesional":     nombre,
                "CUIT":            cuit,
            })
            i += 15
            continue

        # ── FILA DE TOTALIZADOR ───────────────────────────────────────────────
        # Estructura: [cod_ppm][desc][cantidad][anticipo][saldo][mutual][total][gastos][honor]
        if (CODIGO_RE.match(t) and
                i + 7 < n and
                NUM_RE.match(tokens[i + 2] if i + 2 < n else "")):

            def g(o, d=""): return tokens[i + o] if i + o < n else d
            def gn(o):
                try:    return float(g(o).replace(",", "."))
                except: return 0.0

            totalizador.append({
                "Código PPM":      g(0),
                "Descripción PPM": g(1),
                "Cantidad":        gn(2),
                "Anticipo":        gn(3),
                "Saldo":           gn(4),
                "Mutual":          gn(5),
                "Total Facturado": gn(6),
                "Imp Gastos":      gn(7),
                "Imp Honor":       gn(8),
                "Profesional":     nombre,
                "CUIT":            cuit,
            })
            i += 9
            continue

        i += 1

    log.info(f"  → {len(detalle)} filas detalle | {len(totalizador)} totalizador")
    return {"detalle": detalle, "totalizador": totalizador}


def main():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from datetime import datetime

    htmls = list(OUTPUT.rglob("*.html"))
    log.info(f"HTMLs encontrados: {len(htmls)}")
    if not htmls:
        log.error(f"No se encontraron archivos .html en {OUTPUT.resolve()}")
        return

    todas_det = []
    todos_tot = []

    CUIT_RE = re.compile(r"\[?(\d{11})\]?")  # extrae CUIT de 11 dígitos

    for h in htmls:
        nombre = h.stem  # nombre del archivo sin extensión
        cuit   = ""

        # Intentar extraer CUIT desde el contenido del HTML
        try:
            from bs4 import BeautifulSoup
            texto = h.read_bytes()
            for enc in ("utf-8","latin-1","cp1252"):
                try: texto = texto.decode(enc); break
                except: pass
            soup = BeautifulSoup(texto, "lxml")
            tabla = soup.find_all("table")[0]
            tokens_cuit = [td.get_text(" ", strip=True).replace("\xa0"," ").strip()
                           for td in tabla.find_all("td") if td.get_text(strip=True)]
            for tok in tokens_cuit:
                m = CUIT_RE.search(tok)
                if m:
                    cuit = m.group(1)
                    break
        except Exception:
            pass

        datos = parsear_html(h, nombre, cuit)
        todas_det.extend(datos["detalle"])
        todos_tot.extend(datos["totalizador"])

    if not todas_det and not todos_tot:
        log.warning("Sin datos extraíbles. Verificar que los HTML tienen contenido.")
        return

    log.info(f"Total: {len(todas_det)} filas detalle | {len(todos_tot)} totalizador")

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    alt_fill    = PatternFill("solid", fgColor="D6E4F0")
    center      = Alignment(horizontal="center", vertical="center")

    wb = openpyxl.Workbook()

    def make_sheet(ws, headers, filas):
        ws.row_dimensions[1].height = 20
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = header_font; c.fill = header_fill; c.alignment = center
        ws.freeze_panes = "A2"
        for ri, fila in enumerate(filas, 2):
            fill = alt_fill if ri % 2 == 0 else None
            for ci, key in enumerate(headers, 1):
                cell = ws.cell(row=ri, column=ci, value=fila.get(key, ""))
                if fill: cell.fill = fill
        for col in ws.columns:
            ml = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 40)

    # Hoja Detalle
    ws_det = wb.active
    ws_det.title = "Detalle"
    hdrs_det = ["N° Afiliado","Nombre Afiliado","Código PPM","Descripción PPM",
                "Fecha","RRN","Manual","Nro Orden","Cant","Anticipo","Saldo",
                "Mutual","Total Facturado","Imp Gastos","Imp Honor","Terminal",
                "Profesional","CUIT"]
    make_sheet(ws_det, hdrs_det, todas_det)

    # Hoja Totalizador
    ws_tot = wb.create_sheet("Totalizador")
    hdrs_tot = ["Código PPM","Descripción PPM","Cantidad","Anticipo","Saldo",
                "Mutual","Total Facturado","Imp Gastos","Imp Honor","Profesional","CUIT"]
    make_sheet(ws_tot, hdrs_tot, todos_tot)

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    ruta = OUTPUT / f"Consolidado_{MES}_{ANIO}_{ts}.xlsx"
    wb.save(str(ruta))
    log.info(f"\n✓ Excel guardado: {ruta}")
    log.info(f"  Hoja Detalle:     {len(todas_det)} filas")
    log.info(f"  Hoja Totalizador: {len(todos_tot)} filas")


if __name__ == "__main__":
    main()
